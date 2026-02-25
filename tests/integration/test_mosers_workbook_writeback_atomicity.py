from __future__ import annotations

from pathlib import Path

import pytest

from counter_risk.io.mosers_workbook import (
    FuturesDetailNotFoundError,
    atomic_writeback_with_section_locate,
    load_mosers_workbook,
    locate_futures_detail_section,
    writeback_prior_month_notionals,
)


def _build_source_without_futures_detail(path: Path) -> Path:
    from openpyxl import Workbook  # type: ignore[import-untyped]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Inputs"
    worksheet["A1"] = "Not the expected marker"
    worksheet["A2"] = "Description"
    worksheet["B2"] = "Prior Month Notional"
    worksheet["A3"] = "TY Mar25"
    worksheet["B3"] = 123.0
    workbook.save(path)
    workbook.close()
    return path


def test_atomic_writeback_aborts_without_creating_output_when_section_locate_fails(
    tmp_path: Path,
) -> None:
    source_path = _build_source_without_futures_detail(tmp_path / "source.xlsx")
    output_path = tmp_path / "output.xlsx"

    with pytest.raises(FuturesDetailNotFoundError):
        atomic_writeback_with_section_locate(
            source_path=source_path,
            output_path=output_path,
            rows=[{"description": "TY Mar25", "prior_notional": 77.0}],
        )

    assert not output_path.exists()


def test_atomic_writeback_removes_temp_files_when_section_locate_fails(tmp_path: Path) -> None:
    source_path = _build_source_without_futures_detail(tmp_path / "source.xlsx")
    output_path = tmp_path / "output.xlsx"

    with pytest.raises(FuturesDetailNotFoundError):
        atomic_writeback_with_section_locate(
            source_path=source_path,
            output_path=output_path,
            rows=[{"description": "TY Mar25", "prior_notional": 77.0}],
        )

    assert list(tmp_path.glob("*.tmp.xlsx")) == []
    assert list(tmp_path.glob("~$*")) == []


def test_writeback_flow_uses_atomic_wrapper_and_writes_output(tmp_path: Path) -> None:
    source_path = Path("tests/fixtures/mosers_workbook_fixture.xlsx")
    output_path = tmp_path / "output.xlsx"
    expected_description = "US 2-Year Note (CBT) Mar25"
    expected_prior_notional = 77.0

    written_path = writeback_prior_month_notionals(
        source_path=source_path,
        output_path=output_path,
        rows=[{"description": expected_description, "prior_notional": expected_prior_notional}],
    )

    assert written_path == output_path
    assert output_path.exists()

    workbook = load_mosers_workbook(output_path)
    try:
        section = locate_futures_detail_section(workbook)
        worksheet = workbook[section.sheet_name]
        for row_index in range(section.data_start_row, section.data_end_row + 1):
            description = str(
                worksheet.cell(row=row_index, column=section.description_col).value or ""
            ).strip()
            if description == expected_description:
                written_value = worksheet.cell(row=row_index, column=section.prior_month_col).value
                assert written_value == pytest.approx(expected_prior_notional)
                break
        else:
            pytest.fail(f"Could not find expected description row: {expected_description!r}")
    finally:
        workbook.close()
