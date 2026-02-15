"""Tests for MOSERS workbook generation from raw NISA input."""

from __future__ import annotations

from collections.abc import Callable
from pathlib import Path

import pytest

from counter_risk.parsers import parse_cprs_ch, parse_fcm_totals, parse_futures_detail
from counter_risk.writers.mosers_workbook import (
    generate_mosers_workbook,
    generate_mosers_workbook_ex_trend,
    generate_mosers_workbook_trend,
)


def test_generate_mosers_workbook_creates_new_file_with_required_sheets(tmp_path: Path) -> None:
    destination = tmp_path / "MOSERS Generated - All Programs.xlsx"
    output_path = generate_mosers_workbook(
        raw_nisa_path=Path("tests/fixtures/NISA Monthly All Programs - Raw.xlsx"),
        output_path=destination,
    )

    assert output_path == destination
    assert output_path.exists()

    from openpyxl import load_workbook  # type: ignore[import-untyped]

    workbook = load_workbook(output_path, read_only=True, data_only=True)
    try:
        assert "CPRS - CH" in workbook.sheetnames
        assert "CPRS - FCM" in workbook.sheetnames
    finally:
        workbook.close()


def test_generate_mosers_workbook_output_is_parseable_by_existing_milestone_one_parsers(
    tmp_path: Path,
) -> None:
    destination = tmp_path / "MOSERS Generated - All Programs.xlsx"
    output_path = generate_mosers_workbook(
        raw_nisa_path=Path("tests/fixtures/NISA Monthly All Programs - Raw.xlsx"),
        output_path=destination,
    )

    ch_df = parse_cprs_ch(output_path)
    totals_df = parse_fcm_totals(output_path)
    futures_df = parse_futures_detail(output_path)

    assert not ch_df.empty
    assert not totals_df.empty
    assert tuple(totals_df.columns) == (
        "counterparty",
        "TIPS",
        "Treasury",
        "Equity",
        "Commodity",
        "Currency",
        "Notional",
        "NotionalChange",
    )
    assert tuple(futures_df.columns) == (
        "account",
        "description",
        "class",
        "fcm",
        "clearing_house",
        "notional",
    )


@pytest.mark.parametrize(
    (
        "generator",
        "raw_input_name",
        "destination_name",
        "expect_ch_parse",
        "expect_totals_non_empty",
    ),
    [
        (
            generate_mosers_workbook_ex_trend,
            "NISA Monthly Ex Trend - Raw.xlsx",
            "MOSERS Generated - Ex Trend.xlsx",
            True,
            True,
        ),
        (
            generate_mosers_workbook_trend,
            "NISA Monthly Trend - Raw.xlsx",
            "MOSERS Generated - Trend.xlsx",
            False,
            False,
        ),
    ],
)
def test_generate_variant_mosers_workbook_outputs_are_parseable_by_milestone_one_parsers(
    tmp_path: Path,
    generator: Callable[..., Path],
    raw_input_name: str,
    destination_name: str,
    expect_ch_parse: bool,
    expect_totals_non_empty: bool,
) -> None:
    destination = tmp_path / destination_name
    output_path = generator(
        raw_nisa_path=Path("tests/fixtures") / raw_input_name,
        output_path=destination,
    )

    assert output_path == destination
    assert output_path.exists()
    assert output_path.suffix == ".xlsx"

    if expect_ch_parse:
        ch_df = parse_cprs_ch(output_path)
        assert not ch_df.empty

    totals_df = parse_fcm_totals(output_path)
    futures_df = parse_futures_detail(output_path)

    assert totals_df.empty is (not expect_totals_non_empty)
    assert tuple(totals_df.columns) == (
        "counterparty",
        "TIPS",
        "Treasury",
        "Equity",
        "Commodity",
        "Currency",
        "Notional",
        "NotionalChange",
    )
    assert tuple(futures_df.columns) == (
        "account",
        "description",
        "class",
        "fcm",
        "clearing_house",
        "notional",
    )
