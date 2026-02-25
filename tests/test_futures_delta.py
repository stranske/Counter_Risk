"""Additional tests for futures delta: workbook write-back, raw-description sorting,
notional validation, and manifest warning integration.

These tests cover the acceptance criteria introduced in the follow-up PR for issue #29.
"""

from __future__ import annotations

import math
from pathlib import Path
from typing import Any

import pytest

from counter_risk.compute.futures_delta import (
    InvalidNotionalError,
    _extract_notional,
    _validate_row,
    compute_futures_delta,
    normalize_description,
)
from counter_risk.io.mosers_workbook import (
    FuturesDetailNotFoundError,
    FuturesDetailSection,
    load_mosers_workbook,
    locate_futures_detail_section,
    save_mosers_workbook,
    write_prior_month_notional,
)
from counter_risk.pipeline.manifest import WarningsCollector

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_FIXTURE_PATH = Path(__file__).parent / "fixtures" / "mosers_workbook_fixture.xlsx"

_FIXTURE_DESCRIPTIONS = [
    "US 2-Year Note (CBT) Mar25",
    "US 5-Year Note (CBT) Mar25",
    "S&P 500 E-Mini (CME) Mar25",
    "Euro Dollar (CME) Jun25",
    "US Treasury Bond (CBT) Mar25",
]

_PRIOR_NOTIONALS = {
    "US 2-Year Note (CBT) Mar25": 1_200_000.0,
    "US 5-Year Note (CBT) Mar25": 200_000.0,
    "S&P 500 E-Mini (CME) Mar25": 750_000.0,
    "Euro Dollar (CME) Jun25": 0.0,
    "US Treasury Bond (CBT) Mar25": 450_000.0,
}


def _make_rows(*pairs: tuple[str, float]) -> list[dict[str, Any]]:
    return [{"description": desc, "notional": notional} for desc, notional in pairs]


def _records(result: Any) -> list[dict[str, Any]]:
    if hasattr(result, "to_dict"):
        return result.to_dict(orient="records")
    return [dict(row) for row in result]


def _make_result_rows(
    prior_map: dict[str, float] | None = None,
) -> list[dict[str, Any]]:
    """Build result-style rows for the fixture descriptions."""
    if prior_map is None:
        prior_map = _PRIOR_NOTIONALS
    rows = []
    for desc in _FIXTURE_DESCRIPTIONS:
        rows.append(
            {
                "description": desc,
                "notional": 0.0,
                "prior_notional": prior_map.get(desc, 0.0),
                "notional_change": 0.0,
                "sign_flip": "",
            }
        )
    return rows


# ---------------------------------------------------------------------------
# WarningsCollector unit tests
# ---------------------------------------------------------------------------


class TestWarningsCollector:
    def test_empty_on_init(self) -> None:
        col = WarningsCollector()
        assert col.warnings == []

    def test_warn_without_code(self) -> None:
        col = WarningsCollector()
        col.warn("something went wrong")
        assert col.warnings == ["something went wrong"]

    def test_warn_with_code(self) -> None:
        col = WarningsCollector()
        col.warn("bad value", code=WarningsCollector.INVALID_NOTIONAL)
        assert col.warnings == ["[INVALID_NOTIONAL] bad value"]

    def test_warn_multiple(self) -> None:
        col = WarningsCollector()
        col.warn("first", code=WarningsCollector.MISSING_DESCRIPTION)
        col.warn("second")
        col.warn("third", code=WarningsCollector.NO_PRIOR_MONTH_MATCH)
        assert len(col.warnings) == 3

    def test_warnings_returns_copy(self) -> None:
        col = WarningsCollector()
        col.warn("hello")
        w1 = col.warnings
        w1.append("mutate me")
        assert len(col.warnings) == 1  # original unchanged

    def test_warning_codes_are_strings(self) -> None:
        assert isinstance(WarningsCollector.INVALID_NOTIONAL, str)
        assert isinstance(WarningsCollector.MISSING_DESCRIPTION, str)
        assert isinstance(WarningsCollector.MISSING_NOTIONAL, str)
        assert isinstance(WarningsCollector.NO_PRIOR_MONTH_MATCH, str)


# ---------------------------------------------------------------------------
# Workbook write-back tests
# ---------------------------------------------------------------------------


class TestWorkbookWriteBack:
    """Tests for load/locate/write/save cycle using the fixture workbook."""

    def test_load_mosers_workbook(self) -> None:
        wb = load_mosers_workbook(_FIXTURE_PATH)
        assert wb is not None
        assert "Futures Detail" in wb.sheetnames

    def test_load_missing_file_raises(self, tmp_path: Path) -> None:
        with pytest.raises(FileNotFoundError):
            load_mosers_workbook(tmp_path / "nonexistent.xlsx")

    def test_locate_futures_detail_section(self) -> None:
        wb = load_mosers_workbook(_FIXTURE_PATH)
        section = locate_futures_detail_section(wb)
        assert isinstance(section, FuturesDetailSection)
        assert section.sheet_name == "Futures Detail"
        assert section.description_col >= 1
        assert section.prior_month_col >= 1
        assert section.data_start_row > section.header_col_row

    def test_locate_section_finds_correct_columns(self) -> None:
        wb = load_mosers_workbook(_FIXTURE_PATH)
        section = locate_futures_detail_section(wb)
        ws = wb[section.sheet_name]
        # Verify header row has the right column labels
        desc_header = ws.cell(row=section.header_col_row, column=section.description_col).value
        prior_header = ws.cell(row=section.header_col_row, column=section.prior_month_col).value
        assert str(desc_header).strip().lower() == "description"
        assert "prior" in str(prior_header).strip().lower()

    def test_locate_section_data_rows_count(self) -> None:
        wb = load_mosers_workbook(_FIXTURE_PATH)
        section = locate_futures_detail_section(wb)
        n_data_rows = section.data_end_row - section.data_start_row + 1
        assert n_data_rows == len(_FIXTURE_DESCRIPTIONS)

    def test_locate_section_raises_when_no_marker(self, tmp_path: Path) -> None:
        """Workbook with no futures detail marker raises FuturesDetailNotFoundError."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "No marker here"
        path = tmp_path / "no_marker.xlsx"
        wb.save(path)

        wb2 = load_mosers_workbook(path)
        with pytest.raises(FuturesDetailNotFoundError):
            locate_futures_detail_section(wb2)

    def test_locate_section_raises_when_missing_prior_month_col(self, tmp_path: Path) -> None:
        """Missing 'Prior Month Notional' column raises FuturesDetailNotFoundError."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Futures Detail"
        ws["A2"] = "Description"
        ws["B2"] = "Current Month Notional"
        # No "Prior Month Notional" column
        ws["A3"] = "Contract A Mar25"
        path = tmp_path / "missing_col.xlsx"
        wb.save(path)

        wb2 = load_mosers_workbook(path)
        with pytest.raises(FuturesDetailNotFoundError, match="Prior Month Notional"):
            locate_futures_detail_section(wb2)

    def test_write_prior_month_notional_populates_cells(self, tmp_path: Path) -> None:
        """write_prior_month_notional fills prior-month cells for matched rows."""
        wb = load_mosers_workbook(_FIXTURE_PATH)
        section = locate_futures_detail_section(wb)
        rows = _make_result_rows()

        updated = write_prior_month_notional(wb, section, rows)

        assert updated == len(_FIXTURE_DESCRIPTIONS)
        ws = wb[section.sheet_name]
        for data_row in range(section.data_start_row, section.data_end_row + 1):
            desc = str(ws.cell(row=data_row, column=section.description_col).value or "").strip()
            prior_val = ws.cell(row=data_row, column=section.prior_month_col).value
            expected = _PRIOR_NOTIONALS.get(desc)
            assert prior_val == pytest.approx(expected), (
                f"Row {data_row} ({desc!r}): expected {expected}, got {prior_val}"
            )

    def test_write_prior_month_notional_values_are_numeric(self, tmp_path: Path) -> None:
        """Written prior-month values must be numeric (not strings)."""
        wb = load_mosers_workbook(_FIXTURE_PATH)
        section = locate_futures_detail_section(wb)
        write_prior_month_notional(wb, section, _make_result_rows())

        ws = wb[section.sheet_name]
        for data_row in range(section.data_start_row, section.data_end_row + 1):
            prior_val = ws.cell(row=data_row, column=section.prior_month_col).value
            assert isinstance(prior_val, (int, float)), (
                f"Row {data_row}: expected numeric, got {type(prior_val)}"
            )

    def test_write_only_modifies_prior_month_column(self, tmp_path: Path) -> None:
        """Only the prior-month notional column is modified; all other cells unchanged."""
        wb_before = load_mosers_workbook(_FIXTURE_PATH)
        wb_after = load_mosers_workbook(_FIXTURE_PATH)
        section = locate_futures_detail_section(wb_after)
        write_prior_month_notional(wb_after, section, _make_result_rows())

        # Compare every cell across all sheets except the prior-month column of the section.
        for sheet_name in wb_before.sheetnames:
            ws_b = wb_before[sheet_name]
            ws_a = wb_after[sheet_name]
            for row in ws_b.iter_rows():
                for cell in row:
                    r, c = cell.row, cell.column
                    is_prior_col = (
                        sheet_name == section.sheet_name
                        and c == section.prior_month_col
                        and section.data_start_row <= r <= section.data_end_row
                    )
                    if not is_prior_col:
                        after_val = ws_a.cell(row=r, column=c).value
                        assert cell.value == after_val, (
                            f"Sheet {sheet_name!r} cell ({r},{c}): "
                            f"expected {cell.value!r}, got {after_val!r}"
                        )

    def test_save_and_reload_preserves_written_values(self, tmp_path: Path) -> None:
        """Saving and reloading the workbook preserves the written prior-month values."""
        wb = load_mosers_workbook(_FIXTURE_PATH)
        section = locate_futures_detail_section(wb)
        write_prior_month_notional(wb, section, _make_result_rows())

        out_path = tmp_path / "output.xlsx"
        save_mosers_workbook(wb, out_path)
        assert out_path.exists()

        wb2 = load_mosers_workbook(out_path)
        section2 = locate_futures_detail_section(wb2)
        ws2 = wb2[section2.sheet_name]
        for data_row in range(section2.data_start_row, section2.data_end_row + 1):
            desc = str(ws2.cell(row=data_row, column=section2.description_col).value or "")
            prior_val = ws2.cell(row=data_row, column=section2.prior_month_col).value
            expected = _PRIOR_NOTIONALS.get(desc.strip())
            assert prior_val == pytest.approx(expected)

    def test_unmatched_rows_leave_prior_column_unchanged(self, tmp_path: Path) -> None:
        """Rows whose description has no workbook match do not modify any cell."""
        wb = load_mosers_workbook(_FIXTURE_PATH)
        section = locate_futures_detail_section(wb)

        # Provide rows that don't match any workbook description
        no_match_rows = [{"description": "NONEXISTENT Contract Dec25", "prior_notional": 99999.0}]
        updated = write_prior_month_notional(wb, section, no_match_rows)
        assert updated == 0

        ws = wb[section.sheet_name]
        for data_row in range(section.data_start_row, section.data_end_row + 1):
            prior_val = ws.cell(row=data_row, column=section.prior_month_col).value
            assert prior_val is None  # unchanged (fixture has None/blank in prior column)

    def test_partial_match_only_updates_matched_rows(self, tmp_path: Path) -> None:
        """When only a subset of workbook rows are matched, only those are updated."""
        wb = load_mosers_workbook(_FIXTURE_PATH)
        section = locate_futures_detail_section(wb)

        # Only match the first two fixture descriptions
        partial_rows = [
            {"description": _FIXTURE_DESCRIPTIONS[0], "prior_notional": 111.0},
            {"description": _FIXTURE_DESCRIPTIONS[1], "prior_notional": 222.0},
        ]
        updated = write_prior_month_notional(wb, section, partial_rows)
        assert updated == 2

        ws = wb[section.sheet_name]
        # Check that rows 3-5 (0-indexed) still have None in the prior column
        for data_row in range(section.data_start_row + 2, section.data_end_row + 1):
            prior_val = ws.cell(row=data_row, column=section.prior_month_col).value
            assert prior_val is None


# ---------------------------------------------------------------------------
# Raw description sorting tests
# ---------------------------------------------------------------------------


class TestRawDescriptionSorting:
    """Output rows must be sorted by the raw Description text, not normalised."""

    def test_sorted_by_raw_string(self) -> None:
        current = _make_rows(
            ("TY Mar25", 10.0),
            ("ES Mar25", 20.0),
            ("CL Mar25", 30.0),
        )
        prior = _make_rows(
            ("TY Mar25", 5.0),
            ("ES Mar25", 15.0),
            ("CL Mar25", 25.0),
        )
        result = compute_futures_delta(current, prior)
        rows = _records(result)
        raw_descs = [r["description"] for r in rows]
        assert raw_descs == sorted(raw_descs)

    def test_sort_does_not_call_normalize(self) -> None:
        """Verify the sort key is the raw description, not normalize_description().

        We use descriptions where raw sort != normalized sort to detect if
        normalization is inadvertently applied during sorting.
        """
        # "Z-Bond Mar25" raw-sorts AFTER "a-Note Mar25" because 'Z' (90) < 'a' (97)
        # is FALSE in Python (Z=90, a=97 → 'Z' < 'a' is TRUE for ASCII/Unicode).
        # After normalisation both become uppercase: "Z-BOND MAR25" vs "A-NOTE MAR25".
        # Normalized: "A-NOTE MAR25" < "Z-BOND MAR25"
        # Raw: "Z-Bond Mar25" < "a-Note Mar25" (Z=90 < a=97 is TRUE)
        current = _make_rows(
            ("a-Note Mar25", 1.0),
            ("Z-Bond Mar25", 2.0),
        )
        prior: list[dict[str, Any]] = []
        result = compute_futures_delta(current, prior)
        rows = _records(result)
        raw_descs = [r["description"] for r in rows]
        # Raw sort: "Z-Bond Mar25" (Z=90) < "a-Note Mar25" (a=97)
        assert raw_descs == sorted(raw_descs)
        # Normalized sort would give opposite order: "A-NOTE" < "Z-BOND"
        norm_descs = [normalize_description(d) for d in raw_descs]
        # The normalized order differs: verify they are NOT equal to normalized-sorted order
        # (if they were equal it would mean normalization was accidentally used as sort key)
        # For these specific inputs, normalized order is reversed vs raw order.
        assert norm_descs != sorted(norm_descs), (
            "Expected raw sort ≠ normalized sort for these descriptions; "
            "the sort key may be using normalization"
        )

    def test_stable_sort_preserves_duplicate_order(self) -> None:
        """Rows with identical descriptions preserve their original input order."""
        current = [
            {"description": "ES Mar25", "notional": 10.0},
            {"description": "ES Mar25", "notional": 20.0},
            {"description": "ES Mar25", "notional": 30.0},
        ]
        prior: list[dict[str, Any]] = []
        result = compute_futures_delta(current, prior)
        rows = _records(result)
        assert [r["notional"] for r in rows] == pytest.approx([10.0, 20.0, 30.0])

    def test_mixed_descriptions_sorted_correctly(self) -> None:
        current = _make_rows(
            ("US 5-Year Note Mar25", 5.0),
            ("S&P 500 E-Mini Mar25", 500.0),
            ("US 2-Year Note Mar25", 2.0),
            ("Euro Dollar Jun25", 100.0),
        )
        prior: list[dict[str, Any]] = []
        result = compute_futures_delta(current, prior)
        rows = _records(result)
        raw_descs = [r["description"] for r in rows]
        assert raw_descs == sorted(raw_descs)


# ---------------------------------------------------------------------------
# Notional validation tests
# ---------------------------------------------------------------------------


class TestNotionalValidation:
    """Tests for _extract_notional warning emission and strict mode."""

    def test_valid_notional_returns_float(self) -> None:
        row = {"notional": "1234.5"}
        col = WarningsCollector()
        assert _extract_notional(row, row_id="test", collector=col) == pytest.approx(1234.5)
        assert col.warnings == []

    def test_valid_notional_integer(self) -> None:
        row = {"notional": 500}
        assert _extract_notional(row, row_id="test") == pytest.approx(500.0)

    def test_missing_notional_key_emits_warning(self) -> None:
        row = {"description": "ES Mar25"}  # no notional key
        col = WarningsCollector()
        result = _extract_notional(row, row_id="ES Mar25", collector=col)
        assert result == pytest.approx(0.0)
        assert len(col.warnings) == 1
        assert "MISSING_NOTIONAL" in col.warnings[0] or "missing" in col.warnings[0].lower()

    def test_blank_notional_emits_warning(self) -> None:
        row = {"notional": "   "}
        col = WarningsCollector()
        result = _extract_notional(row, row_id="Row 0", collector=col)
        assert result == pytest.approx(0.0)
        assert len(col.warnings) == 1
        assert "INVALID_NOTIONAL" in col.warnings[0]

    def test_non_numeric_notional_emits_warning(self) -> None:
        row = {"notional": "not-a-number"}
        col = WarningsCollector()
        result = _extract_notional(row, row_id="Row 1", collector=col)
        assert result == pytest.approx(0.0)
        assert len(col.warnings) == 1
        assert "INVALID_NOTIONAL" in col.warnings[0]

    def test_nan_notional_emits_warning(self) -> None:
        row = {"notional": float("nan")}
        col = WarningsCollector()
        result = _extract_notional(row, row_id="Row 2", collector=col)
        assert result == pytest.approx(0.0)
        assert len(col.warnings) == 1
        assert "INVALID_NOTIONAL" in col.warnings[0]

    def test_nan_as_math_nan_emits_warning(self) -> None:
        row = {"notional": math.nan}
        col = WarningsCollector()
        result = _extract_notional(row, row_id="Row 3", collector=col)
        assert result == pytest.approx(0.0)
        assert len(col.warnings) == 1

    def test_strict_mode_raises_on_blank_notional(self) -> None:
        row = {"notional": ""}
        col = WarningsCollector()
        with pytest.raises(InvalidNotionalError, match="Row 0"):
            _extract_notional(row, row_id="Row 0", strict=True, collector=col)

    def test_strict_mode_raises_on_non_numeric(self) -> None:
        row = {"notional": "abc"}
        col = WarningsCollector()
        with pytest.raises(InvalidNotionalError, match="abc"):
            _extract_notional(row, row_id="Row 5", strict=True, collector=col)

    def test_strict_mode_raises_on_nan(self) -> None:
        row = {"notional": float("nan")}
        col = WarningsCollector()
        with pytest.raises(InvalidNotionalError):
            _extract_notional(row, row_id="Row 6", strict=True, collector=col)

    def test_strict_mode_raises_on_missing_key(self) -> None:
        row = {"description": "ES Mar25"}
        col = WarningsCollector()
        with pytest.raises(InvalidNotionalError, match="[Mm]issing"):
            _extract_notional(row, row_id="ES Mar25", strict=True, collector=col)

    def test_strict_error_includes_row_identifier(self) -> None:
        row = {"notional": "bad"}
        with pytest.raises(InvalidNotionalError) as exc_info:
            _extract_notional(row, row_id="my-row-identifier", strict=True)
        assert "my-row-identifier" in str(exc_info.value)

    def test_notional_alias_capitalized(self) -> None:
        row = {"Notional": 999.0}
        assert _extract_notional(row) == pytest.approx(999.0)

    def test_notional_alias_exposure(self) -> None:
        row = {"exposure": 777.0}
        assert _extract_notional(row) == pytest.approx(777.0)

    def test_warning_not_silent_for_invalid(self) -> None:
        """Missing/invalid notional must emit a warning, not silently return 0.0."""
        row = {"notional": "bad"}
        col = WarningsCollector()
        _extract_notional(row, row_id="Row X", collector=col)
        # A warning must have been emitted (not silently returning 0.0)
        assert len(col.warnings) > 0


# ---------------------------------------------------------------------------
# Per-row required-field validation tests
# ---------------------------------------------------------------------------


class TestValidateRow:
    """Tests for _validate_row field validation."""

    def test_valid_row_returns_true(self) -> None:
        row = {"description": "ES Mar25", "notional": 100.0}
        col = WarningsCollector()
        assert _validate_row(row, row_idx=0, collector=col) is True
        assert col.warnings == []

    def test_missing_description_returns_false(self) -> None:
        row = {"notional": 100.0}
        col = WarningsCollector()
        assert _validate_row(row, row_idx=0, collector=col) is False
        assert len(col.warnings) == 1
        assert "MISSING_DESCRIPTION" in col.warnings[0]

    def test_blank_description_returns_false(self) -> None:
        row = {"description": "   ", "notional": 100.0}
        col = WarningsCollector()
        assert _validate_row(row, row_idx=0, collector=col) is False
        assert "MISSING_DESCRIPTION" in col.warnings[0]

    def test_missing_notional_returns_false(self) -> None:
        row = {"description": "ES Mar25"}  # no notional key
        col = WarningsCollector()
        assert _validate_row(row, row_idx=0, collector=col) is False
        assert len(col.warnings) == 1
        assert "MISSING_NOTIONAL" in col.warnings[0]

    def test_invalid_description_warning_includes_row_idx(self) -> None:
        row = {"description": "", "notional": 100.0}
        col = WarningsCollector()
        _validate_row(row, row_idx=7, collector=col)
        assert "7" in col.warnings[0]

    def test_invalid_description_warning_includes_non_empty_fields(self) -> None:
        row = {"description": "", "notional": 42.0, "other": "x"}
        col = WarningsCollector()
        _validate_row(row, row_idx=0, collector=col)
        # Warning should include available non-empty field values
        assert "notional" in col.warnings[0] or "42" in col.warnings[0]

    def test_invalid_row_excluded_from_compute(self) -> None:
        """Rows with missing/blank Description are excluded from delta computation."""
        current = [
            {"description": "", "notional": 999.0},  # invalid - no description
            {"description": "ES Mar25", "notional": 100.0},  # valid
        ]
        prior: list[dict[str, Any]] = []
        col = WarningsCollector()
        result = compute_futures_delta(current, prior, collector=col)
        rows = _records(result)
        # Only the valid row should appear in output
        assert len(rows) == 1
        assert rows[0]["description"] == "ES Mar25"
        # A warning should have been emitted for the invalid row
        assert any("MISSING_DESCRIPTION" in w for w in col.warnings)

    def test_row_with_missing_notional_excluded(self) -> None:
        """Rows with no notional key are excluded from delta computation."""
        current = [
            {"description": "No Notional Row"},  # missing notional key
            {"description": "ES Mar25", "notional": 50.0},
        ]
        prior: list[dict[str, Any]] = []
        col = WarningsCollector()
        result = compute_futures_delta(current, prior, collector=col)
        rows = _records(result)
        assert len(rows) == 1
        assert rows[0]["description"] == "ES Mar25"
        assert any("MISSING_NOTIONAL" in w for w in col.warnings)

    def test_description_key_case_insensitive(self) -> None:
        """Both 'description' and 'Description' keys are accepted."""
        row_lower = {"description": "ES Mar25", "notional": 1.0}
        row_upper = {"Description": "ES Mar25", "notional": 1.0}
        assert _validate_row(row_lower, row_idx=0) is True
        assert _validate_row(row_upper, row_idx=0) is True


# ---------------------------------------------------------------------------
# Manifest/collector integration tests for unmatched rows
# ---------------------------------------------------------------------------


class TestUnmatchedRowManifestWarnings:
    """Unmatched rows are reported via WarningsCollector with correct reason codes."""

    def test_unmatched_current_emits_no_prior_match_code(self) -> None:
        current = _make_rows(("NEW Contract Dec25", 100.0))
        prior: list[dict[str, Any]] = []
        col = WarningsCollector()
        compute_futures_delta(current, prior, collector=col)
        # Exactly one warning for the unmatched current row
        unmatched = [w for w in col.warnings if "Unmatched current" in w]
        assert len(unmatched) == 1
        assert "NO_PRIOR_MONTH_MATCH" in unmatched[0]

    def test_unmatched_current_warning_includes_description(self) -> None:
        current = _make_rows(("MY UNIQUE TICKER Mar25", 1.0))
        prior: list[dict[str, Any]] = []
        col = WarningsCollector()
        compute_futures_delta(current, prior, collector=col)
        assert any("MY UNIQUE TICKER" in w for w in col.warnings)

    def test_unmatched_prior_emits_no_prior_match_code(self) -> None:
        current: list[dict[str, Any]] = []
        prior = _make_rows(("Old Contract Dec25", 50.0))
        col = WarningsCollector()
        compute_futures_delta(current, prior, collector=col)
        unmatched = [w for w in col.warnings if "Unmatched prior" in w]
        assert len(unmatched) == 1
        assert "NO_PRIOR_MONTH_MATCH" in unmatched[0]

    def test_one_warning_per_unmatched_row(self) -> None:
        current = _make_rows(
            ("Contract A Mar25", 10.0),
            ("Contract B Mar25", 20.0),
            ("Contract C Mar25", 30.0),
        )
        prior: list[dict[str, Any]] = []
        col = WarningsCollector()
        compute_futures_delta(current, prior, collector=col)
        unmatched = [w for w in col.warnings if "Unmatched current" in w]
        assert len(unmatched) == 3

    def test_matched_rows_produce_no_unmatched_warning(self) -> None:
        current = _make_rows(("ES Mar25", 100.0))
        prior = _make_rows(("ES Mar25", 80.0))
        col = WarningsCollector()
        compute_futures_delta(current, prior, collector=col)
        assert col.warnings == []

    def test_unmatched_warning_not_returned_in_function_result(self) -> None:
        """compute_futures_delta returns only the result table, not a warnings tuple."""
        current = _make_rows(("Unmatched Mar25", 1.0))
        prior: list[dict[str, Any]] = []
        result = compute_futures_delta(current, prior)
        # Result must NOT be a tuple (no longer (result, warnings) pattern).
        assert not isinstance(result, tuple)

    def test_no_collector_still_works(self) -> None:
        """Calling without a collector should not raise; warnings go to logger only."""
        current = _make_rows(("ES Mar25", 100.0))
        prior: list[dict[str, Any]] = []
        result = compute_futures_delta(current, prior)
        rows = _records(result)
        assert len(rows) == 1
