"""Tests for the Excel Runner workbook artifact."""

from __future__ import annotations

import hashlib
import re
from datetime import date
from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZipFile

import pytest
from scripts import build_runner_workbook as runner_builder

SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
RELATIONSHIP_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
MARKUP_COMPATIBILITY_NS = "http://schemas.openxmlformats.org/markup-compatibility/2006"
SPREADSHEET_DRAWING_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
VML_NS = "urn:schemas-microsoft-com:vml"
EXCEL_NS = "urn:schemas-microsoft-com:office:excel"
X14_NS = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"

CTRL_PROP_CONTENT_TYPE = "application/vnd.ms-excel.controlproperties+xml"
CTRL_PROP_REL_TYPE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/ctrlProp"
VML_DRAWING_REL_TYPE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing"
)

NAMESPACES = {
    "ss": SPREADSHEET_NS,
    "r": RELATIONSHIP_NS,
    "pr": PACKAGE_REL_NS,
    "ct": CONTENT_TYPES_NS,
    "mc": MARKUP_COMPATIBILITY_NS,
    "xdr": SPREADSHEET_DRAWING_NS,
    "v": VML_NS,
    "x": EXCEL_NS,
    "x14": X14_NS,
}

EXPECTED_RUNNER_ACTION_CELLS = {
    "A5": "Run All",
    "B5": "Run Ex Trend",
    "C5": "Run Trend",
    "D5": "Dry-Run Discovery",
    "E5": "Open Output Folder",
    "F5": "Open Manifest",
    "G5": "Open Summary",
    "H5": "Open PPT Folder",
    "I5": "Ask about this run",
}

EXPECTED_RUNNER_BUTTONS = {
    "A5": ("Run All", "RunAll_Click", 1025, "rId2", "ctrlProp1.xml"),
    "B5": ("Run Ex Trend", "RunExTrend_Click", 1026, "rId3", "ctrlProp2.xml"),
    "C5": ("Run Trend", "RunTrend_Click", 1027, "rId4", "ctrlProp3.xml"),
    "E5": ("Open Output Folder", "OpenOutputFolder_Click", 1028, "rId5", "ctrlProp4.xml"),
    "F5": ("Open Manifest", "OpenManifest_Click", 1029, "rId6", "ctrlProp5.xml"),
    "G5": ("Open Summary", "OpenSummary_Click", 1030, "rId7", "ctrlProp6.xml"),
    "H5": ("Open PPT Folder", "OpenPPTFolder_Click", 1031, "rId8", "ctrlProp7.xml"),
}
PUBLIC_SUB_PATTERN = re.compile(r"^Public Sub ([A-Za-z0-9_]+)\(\)", re.MULTILINE)


def _is_month_end(iso_date: str) -> bool:
    parsed = date.fromisoformat(iso_date)
    next_day = parsed.fromordinal(parsed.toordinal() + 1)
    return next_day.month != parsed.month


def _next_month_end(current: date) -> date:
    next_month_start = (
        date(current.year + 1, 1, 1)
        if current.month == 12
        else date(current.year, current.month + 1, 1)
    )
    month_after_next_start = (
        date(next_month_start.year + 1, 1, 1)
        if next_month_start.month == 12
        else date(next_month_start.year, next_month_start.month + 1, 1)
    )
    return month_after_next_start.fromordinal(month_after_next_start.toordinal() - 1)


def _read_xml(zip_file: ZipFile, member: str) -> ElementTree.Element:
    with zip_file.open(member) as handle:
        return ElementTree.fromstring(handle.read())  # noqa: S314


def _sha256_bytes(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _macro_name(macro_formula: str) -> str:
    return macro_formula.rsplit("!", maxsplit=1)[-1]


def _extract_vba_project_bytes(workbook_path: Path) -> bytes:
    with ZipFile(workbook_path) as zip_file, zip_file.open("xl/vbaProject.bin") as handle:
        return handle.read()


def _extract_searchable_vba_text(workbook_path: Path) -> str:
    vba_bytes = _extract_vba_project_bytes(workbook_path)
    try:
        return vba_bytes.decode("latin-1", errors="ignore")
    except Exception:  # pragma: no cover - defensive fallback for non-standard decode failures
        return vba_bytes.decode("latin-1", errors="replace")


def test_runner_workbook_exists() -> None:
    assert Path("Runner.xlsm").is_file(), "Runner.xlsm must be committed to the repository root."


def test_runner_workbook_has_required_ooxml_structure_and_content_types() -> None:
    workbook_path = Path("Runner.xlsm")
    with ZipFile(workbook_path) as zip_file:
        members = set(zip_file.namelist())
        expected_members = {
            "[Content_Types].xml",
            "_rels/.rels",
            "docProps/app.xml",
            "docProps/core.xml",
            "xl/workbook.xml",
            "xl/_rels/workbook.xml.rels",
            "xl/styles.xml",
            "xl/vbaProject.bin",
            "xl/worksheets/sheet1.xml",
            "xl/worksheets/_rels/sheet1.xml.rels",
            "xl/worksheets/sheet2.xml",
            "xl/worksheets/sheet3.xml",
            "xl/worksheets/sheet4.xml",
            "xl/drawings/vmlDrawing1.vml",
        }
        expected_members.update(
            f"xl/ctrlProps/{ctrl_prop_name}"
            for *_, ctrl_prop_name in EXPECTED_RUNNER_BUTTONS.values()
        )
        assert expected_members.issubset(members)

        content_types_root = _read_xml(zip_file, "[Content_Types].xml")
        defaults = content_types_root.findall("ct:Default", NAMESPACES)
        default_content_types = {
            node.attrib["Extension"]: node.attrib["ContentType"] for node in defaults
        }
        overrides = content_types_root.findall("ct:Override", NAMESPACES)
        override_content_types = {
            node.attrib["PartName"]: node.attrib["ContentType"] for node in overrides
        }

        assert (
            default_content_types["vml"]
            == "application/vnd.openxmlformats-officedocument.vmlDrawing"
        )
        assert (
            override_content_types["/xl/workbook.xml"]
            == "application/vnd.ms-excel.sheet.macroEnabled.main+xml"
        )
        assert (
            override_content_types["/xl/vbaProject.bin"] == "application/vnd.ms-office.vbaProject"
        )
        assert (
            override_content_types["/xl/worksheets/sheet1.xml"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"
        )
        assert (
            override_content_types["/xl/worksheets/sheet2.xml"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"
        )
        assert (
            override_content_types["/xl/worksheets/sheet3.xml"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"
        )
        assert (
            override_content_types["/xl/worksheets/sheet4.xml"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"
        )
        assert (
            override_content_types["/xl/styles.xml"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"
        )
        for *_, ctrl_prop_name in EXPECTED_RUNNER_BUTTONS.values():
            assert (
                override_content_types[f"/xl/ctrlProps/{ctrl_prop_name}"] == CTRL_PROP_CONTENT_TYPE
            )

        with zip_file.open("xl/vbaProject.bin") as handle:
            vba_project = handle.read()
        assert len(vba_project) >= 1024

        workbook_rels_root = _read_xml(zip_file, "xl/_rels/workbook.xml.rels")
        rels = workbook_rels_root.findall("pr:Relationship", NAMESPACES)
        vba_project_rel = next(
            (
                rel
                for rel in rels
                if rel.attrib.get("Type")
                == "http://schemas.microsoft.com/office/2006/relationships/vbaProject"
            ),
            None,
        )
        assert vba_project_rel is not None
        assert vba_project_rel.attrib.get("Target") == "vbaProject.bin"


def test_runner_workbook_contains_month_selector_dropdown() -> None:
    workbook_path = Path("Runner.xlsm")
    with ZipFile(workbook_path) as zip_file:
        workbook_root = _read_xml(zip_file, "xl/workbook.xml")
        workbook_rels_root = _read_xml(zip_file, "xl/_rels/workbook.xml.rels")

        sheets = workbook_root.findall("ss:sheets/ss:sheet", NAMESPACES)
        sheet_by_name = {sheet.attrib["name"]: sheet for sheet in sheets}

        assert "Runner" in sheet_by_name
        assert "ControlData" in sheet_by_name
        assert "Settings" in sheet_by_name
        assert "Config" in sheet_by_name
        assert sheet_by_name["ControlData"].attrib.get("state") == "hidden"

        rel_target_by_id = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in workbook_rels_root.findall("pr:Relationship", NAMESPACES)
        }

        runner_sheet_target = rel_target_by_id[
            sheet_by_name["Runner"].attrib[f"{{{RELATIONSHIP_NS}}}id"]
        ]
        control_sheet_target = rel_target_by_id[
            sheet_by_name["ControlData"].attrib[f"{{{RELATIONSHIP_NS}}}id"]
        ]
        settings_sheet_target = rel_target_by_id[
            sheet_by_name["Settings"].attrib[f"{{{RELATIONSHIP_NS}}}id"]
        ]
        config_sheet_target = rel_target_by_id[
            sheet_by_name["Config"].attrib[f"{{{RELATIONSHIP_NS}}}id"]
        ]

        runner_sheet_root = _read_xml(zip_file, f"xl/{runner_sheet_target}")
        control_sheet_root = _read_xml(zip_file, f"xl/{control_sheet_target}")
        settings_sheet_root = _read_xml(zip_file, f"xl/{settings_sheet_target}")
        config_sheet_root = _read_xml(zip_file, f"xl/{config_sheet_target}")

        selector_label = runner_sheet_root.find(
            "ss:sheetData/ss:row[@r='3']/ss:c[@r='A3']/ss:is/ss:t",
            NAMESPACES,
        )
        assert selector_label is not None
        assert selector_label.text == "As-Of Month"

        validations = runner_sheet_root.findall("ss:dataValidations/ss:dataValidation", NAMESPACES)
        assert validations, "Runner sheet must include data validation on the month selector cell."

        month_validation = next(
            (item for item in validations if item.attrib.get("sqref") == "B3"), None
        )
        assert month_validation is not None
        assert month_validation.attrib.get("type") == "list"

        formula = month_validation.find("ss:formula1", NAMESPACES)
        assert formula is not None
        assert formula.text == "ControlData!$A$2:$A$193"

        month_cells = control_sheet_root.findall("ss:sheetData/ss:row/ss:c/ss:is/ss:t", NAMESPACES)
        month_values = [cell.text for cell in month_cells if cell.text is not None]

        assert month_values[0] == "MonthEnd"
        assert month_values[1] == "2020-01-31"
        assert month_values[-1] == "2035-12-31"
        assert all(_is_month_end(item) for item in month_values[1:])

        parsed_dates = [date.fromisoformat(item) for item in month_values[1:]]
        for current, next_value in zip(parsed_dates, parsed_dates[1:], strict=False):
            assert next_value == _next_month_end(current)

        settings_labels = settings_sheet_root.findall(
            "ss:sheetData/ss:row/ss:c/ss:is/ss:t",
            NAMESPACES,
        )
        settings_values = [node.text for node in settings_labels if node.text]
        assert "Discovery Mode" in settings_values
        assert "Strict Policy" in settings_values
        assert "Formatting Profile" in settings_values
        assert "Output Root" in settings_values

        config_labels = config_sheet_root.findall(
            "ss:sheetData/ss:row/ss:c/ss:is/ss:t",
            NAMESPACES,
        )
        config_values = [node.text for node in config_labels if node.text]
        assert "MOSERS All Programs (.xlsx)" in config_values
        assert "MOSERS Ex-Trend (.xlsx)" in config_values
        assert "Monthly Report Template (.pptx)" in config_values


def test_runner_workbook_contains_run_controls() -> None:
    workbook_path = Path("Runner.xlsm")
    with ZipFile(workbook_path) as zip_file:
        workbook_root = _read_xml(zip_file, "xl/workbook.xml")
        workbook_rels_root = _read_xml(zip_file, "xl/_rels/workbook.xml.rels")

        sheets = workbook_root.findall("ss:sheets/ss:sheet", NAMESPACES)
        sheet_by_name = {sheet.attrib["name"]: sheet for sheet in sheets}
        rel_target_by_id = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in workbook_rels_root.findall("pr:Relationship", NAMESPACES)
        }

        runner_sheet_target = rel_target_by_id[
            sheet_by_name["Runner"].attrib[f"{{{RELATIONSHIP_NS}}}id"]
        ]
        runner_sheet_root = _read_xml(zip_file, f"xl/{runner_sheet_target}")

        for cell_ref, expected_text in EXPECTED_RUNNER_ACTION_CELLS.items():
            node = runner_sheet_root.find(
                f"ss:sheetData/ss:row[@r='5']/ss:c[@r='{cell_ref}']/ss:is/ss:t",
                NAMESPACES,
            )
            assert node is not None
            assert node.text == expected_text

        legacy_drawing = runner_sheet_root.find("ss:legacyDrawing", NAMESPACES)
        assert legacy_drawing is not None
        assert legacy_drawing.attrib[f"{{{RELATIONSHIP_NS}}}id"] == "rId1"

        controls = runner_sheet_root.findall(
            "ss:controls/mc:AlternateContent/mc:Choice/ss:control",
            NAMESPACES,
        )
        assert len(controls) == len(EXPECTED_RUNNER_BUTTONS)
        controls_by_caption = {
            control.attrib["name"].removesuffix(" Button"): control for control in controls
        }

        sheet_rels_root = _read_xml(zip_file, "xl/worksheets/_rels/sheet1.xml.rels")
        rels_by_id = {
            rel.attrib["Id"]: rel.attrib
            for rel in sheet_rels_root.findall("pr:Relationship", NAMESPACES)
        }
        assert rels_by_id["rId1"]["Type"] == VML_DRAWING_REL_TYPE
        assert rels_by_id["rId1"]["Target"] == "../drawings/vmlDrawing1.vml"

        vml_root = _read_xml(zip_file, "xl/drawings/vmlDrawing1.vml")
        vml_shapes = vml_root.findall("v:shape", NAMESPACES)
        assert len(vml_shapes) == len(EXPECTED_RUNNER_BUTTONS)
        vml_shape_by_id = {shape.attrib["id"]: shape for shape in vml_shapes}

        for cell_ref, (
            caption,
            macro,
            shape_id,
            relationship_id,
            ctrl_prop_name,
        ) in EXPECTED_RUNNER_BUTTONS.items():
            control = controls_by_caption[caption]
            assert control.attrib["shapeId"] == str(shape_id)
            assert control.attrib[f"{{{RELATIONSHIP_NS}}}id"] == relationship_id

            control_properties = control.find("ss:controlPr", NAMESPACES)
            assert control_properties is not None
            assert _macro_name(control_properties.attrib["macro"]) == macro
            assert control_properties.attrib["altText"] == caption

            from_marker = control_properties.find("ss:anchor/xdr:from", NAMESPACES)
            assert from_marker is not None
            expected_column = ord(cell_ref[0]) - ord("A")
            assert from_marker.find("xdr:col", NAMESPACES).text == str(expected_column)
            assert from_marker.find("xdr:row", NAMESPACES).text == "4"

            ctrl_prop_rel = rels_by_id[relationship_id]
            assert ctrl_prop_rel["Type"] == CTRL_PROP_REL_TYPE
            assert ctrl_prop_rel["Target"] == f"../ctrlProps/{ctrl_prop_name}"
            ctrl_prop_root = _read_xml(zip_file, f"xl/ctrlProps/{ctrl_prop_name}")
            assert ctrl_prop_root.tag == f"{{{X14_NS}}}formControlPr"
            assert ctrl_prop_root.attrib["objectType"] == "Button"
            assert ctrl_prop_root.attrib["textHAlign"] == "center"
            assert ctrl_prop_root.attrib["textVAlign"] == "center"

            shape = vml_shape_by_id[f"_x0000_s{shape_id}"]
            assert shape.attrib["{urn:schemas-microsoft-com:office:office}button"] == "t"
            vml_caption = shape.find(".//font", NAMESPACES)
            assert vml_caption is not None
            assert vml_caption.text == caption
            vml_macro = shape.find("x:ClientData/x:FmlaMacro", NAMESPACES)
            assert vml_macro is not None
            assert _macro_name(vml_macro.text) == macro


def test_runner_workbook_button_macros_exist_in_runnerlaunch_source() -> None:
    source = Path("assets/vba/RunnerLaunch.bas").read_text(encoding="utf-8")
    public_subs = set(PUBLIC_SUB_PATTERN.findall(source))

    workbook_path = Path("Runner.xlsm")
    with ZipFile(workbook_path) as zip_file:
        vml_root = _read_xml(zip_file, "xl/drawings/vmlDrawing1.vml")

    macros = {
        _macro_name(node.text)
        for node in vml_root.findall("v:shape/x:ClientData/x:FmlaMacro", NAMESPACES)
        if node.text
    }

    assert macros
    assert macros <= public_subs


def test_runner_workbook_loads_with_openpyxl_keep_vba() -> None:
    from openpyxl import load_workbook

    workbook = load_workbook("Runner.xlsm", keep_vba=True)

    assert workbook.sheetnames == ["Runner", "ControlData", "Settings", "Config"]
    assert workbook.vba_archive is not None


def test_runner_workbook_defines_named_ranges_for_settings_cells() -> None:
    workbook_path = Path("Runner.xlsm")
    with ZipFile(workbook_path) as zip_file:
        workbook_root = _read_xml(zip_file, "xl/workbook.xml")
        defined_names = workbook_root.findall("ss:definedNames/ss:definedName", NAMESPACES)
        names = {node.attrib.get("name"): node.text for node in defined_names}

    assert names["RunnerSetting_InputRoot"] == "Settings!$B$2"
    assert names["RunnerSetting_DiscoveryMode"] == "Settings!$B$3"
    assert names["RunnerSetting_StrictPolicy"] == "Settings!$B$4"
    assert names["RunnerSetting_FormattingProfile"] == "Settings!$B$5"
    assert names["RunnerSetting_OutputRoot"] == "Settings!$B$6"
    assert names["RunnerConfig_MosersAllPrograms"] == "Config!$B$4"
    assert names["RunnerConfig_MosersExTrend"] == "Config!$B$5"
    assert names["RunnerConfig_MosersTrend"] == "Config!$B$6"
    assert names["RunnerConfig_HistAllPrograms3yr"] == "Config!$B$7"
    assert names["RunnerConfig_HistExLlc3yr"] == "Config!$B$8"
    assert names["RunnerConfig_HistLlc3yr"] == "Config!$B$9"
    assert names["RunnerConfig_MonthlyPptx"] == "Config!$B$10"


def test_runner_workbook_contains_data_quality_row() -> None:
    workbook_path = Path("Runner.xlsm")
    with ZipFile(workbook_path) as zip_file:
        workbook_root = _read_xml(zip_file, "xl/workbook.xml")
        workbook_rels_root = _read_xml(zip_file, "xl/_rels/workbook.xml.rels")

        sheets = workbook_root.findall("ss:sheets/ss:sheet", NAMESPACES)
        sheet_by_name = {sheet.attrib["name"]: sheet for sheet in sheets}
        rel_target_by_id = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in workbook_rels_root.findall("pr:Relationship", NAMESPACES)
        }

        runner_sheet_target = rel_target_by_id[
            sheet_by_name["Runner"].attrib[f"{{{RELATIONSHIP_NS}}}id"]
        ]
        runner_sheet_root = _read_xml(zip_file, f"xl/{runner_sheet_target}")

        label_node = runner_sheet_root.find(
            "ss:sheetData/ss:row[@r='9']/ss:c[@r='A9']/ss:is/ss:t",
            NAMESPACES,
        )
        assert label_node is not None
        assert label_node.text == "Data Quality"

        status_cell = runner_sheet_root.find(
            "ss:sheetData/ss:row[@r='9']/ss:c[@r='B9']",
            NAMESPACES,
        )
        assert status_cell is not None

        status_label = runner_sheet_root.find(
            "ss:sheetData/ss:row[@r='7']/ss:c[@r='A7']/ss:is/ss:t",
            NAMESPACES,
        )
        assert status_label is not None
        assert status_label.text == "Status"

        result_label = runner_sheet_root.find(
            "ss:sheetData/ss:row[@r='8']/ss:c[@r='A8']/ss:is/ss:t",
            NAMESPACES,
        )
        assert result_label is not None
        assert result_label.text == "Result"


def test_build_runner_workbook_embeds_valid_vba_binary_with_matching_hash(tmp_path: Path) -> None:
    output_workbook = tmp_path / "Runner.built.xlsm"
    runner_builder.build_runner_workbook(output_workbook)

    built_vba_project = _extract_vba_project_bytes(output_workbook)
    source_vba_project = Path("assets/vba/vbaProject.bin").read_bytes()

    assert source_vba_project[:8] == runner_builder.OLE_CFB_SIGNATURE
    assert built_vba_project[:8] == runner_builder.OLE_CFB_SIGNATURE
    assert _sha256_bytes(built_vba_project) == _sha256_bytes(source_vba_project)


def test_build_runner_workbook_extracts_vba_binary_with_ole_signature(tmp_path: Path) -> None:
    output_workbook = tmp_path / "Runner.built.xlsm"
    runner_builder.build_runner_workbook(output_workbook)

    extracted_vba_project = _extract_vba_project_bytes(output_workbook)

    assert extracted_vba_project[:8] == runner_builder.OLE_CFB_SIGNATURE


def test_build_runner_workbook_extracts_searchable_vba_text(tmp_path: Path) -> None:
    output_workbook = tmp_path / "Runner.built.xlsm"
    runner_builder.build_runner_workbook(output_workbook)

    vba_text = _extract_searchable_vba_text(output_workbook)

    assert vba_text
    assert 'Attribute VB_Name = "RunnerLaunch"' in vba_text
    assert "Public Sub RunAll_Click()" in vba_text
    assert "Public Sub RunExTrend_Click()" in vba_text
    assert "Public Sub RunTrend_Click()" in vba_text
    assert "OpenOutputFolder_Click" in vba_text
    assert "BuildRunArguments" in vba_text
    assert "BuildExecutableCommand" in vba_text
    assert "ReadSelectedDate" in vba_text
    assert "Running..." in vba_text
    assert "Success" in vba_text
    assert "Error" in vba_text
    assert "Directory not found" in vba_text
    assert 'MsgBox "Directory not found" & " " & p' in vba_text
    assert 'MsgBox "Directory not found" & resolvedPath' not in vba_text


def test_build_runner_workbook_fails_when_vba_project_bin_missing(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    missing_path = tmp_path / "missing-vbaProject.bin"
    monkeypatch.setattr(runner_builder, "VBA_PROJECT_PATH", missing_path)

    with pytest.raises(FileNotFoundError):
        runner_builder.build_runner_workbook(tmp_path / "Runner.built.xlsm")
    assert runner_builder.main() == 1


def test_build_runner_workbook_fails_when_vba_project_bin_has_invalid_signature(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    invalid_path = tmp_path / "invalid-vbaProject.bin"
    invalid_path.write_bytes(b"NOT-OLE-BINARY")
    monkeypatch.setattr(runner_builder, "VBA_PROJECT_PATH", invalid_path)

    with pytest.raises(ValueError):
        runner_builder.build_runner_workbook(tmp_path / "Runner.built.xlsm")
    assert runner_builder.main() == 1
