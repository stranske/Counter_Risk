"""Fixture smoke tests for macro-spec parity coverage."""

from __future__ import annotations

import math
from pathlib import Path
from typing import Any

import pytest

_MACRO_BY_VARIANT = {
    "all_programs": "RunAll_Click",
    "ex_trend": "RunExTrend_Click",
    "trend": "RunTrend_Click",
}
_TEMPLATE_WORKBOOK_PATH = Path(
    "tests/fixtures/MOSERS Counterparty Risk Summary 12-31-2025 - All Programs.xlsx"
)


def test_macro_spec_fixture_inputs_exist(macro_spec_cases: tuple[Any, ...]) -> None:
    for case in macro_spec_cases:
        assert case.raw_input_path.is_file(), (
            f"Missing macro-spec fixture input for {case.variant}: {case.raw_input_path}"
        )


def test_macro_spec_fixture_parsers_load_sample_inputs(
    parsed_macro_spec_data: dict[str, Any],
) -> None:
    for variant, parsed in parsed_macro_spec_data.items():
        assert parsed.ch_rows, f"{variant}: expected at least one CPRS-CH row from fixture input"
        assert parsed.totals_rows, f"{variant}: expected at least one totals row from fixture input"


def test_macro_spec_fixture_generators_create_cprs_ch_sheet(
    macro_spec_cases: tuple[Any, ...],
) -> None:
    for case in macro_spec_cases:
        workbook = case.generator(case.raw_input_path)
        try:
            assert "CPRS - CH" in workbook.sheetnames, (
                f"{case.variant}: generated workbook missing required CPRS - CH sheet"
            )
        finally:
            workbook.close()


def test_macro_spec_range_level_outputs_match_expected_mosers_format(
    macro_spec_cases: tuple[Any, ...],
    parsed_macro_spec_data: dict[str, Any],
) -> None:
    for case in macro_spec_cases:
        parsed = parsed_macro_spec_data[case.variant]
        workbook = case.generator(case.raw_input_path)
        try:
            worksheet = workbook["CPRS - CH"]
            macro_name = _MACRO_BY_VARIANT[case.variant]

            expected_vols = _pad_to_slot_count(
                [row.annualized_volatility for row in parsed.totals_rows], 11
            )
            expected_allocations = _pad_to_slot_count(_expected_allocations(parsed), 11)

            actual_vols = [worksheet[f"D{row_number}"].value for row_number in range(10, 21)]
            actual_allocations = [worksheet[f"E{row_number}"].value for row_number in range(10, 21)]

            assert worksheet["B5"].value == parsed.ch_rows[0].counterparty, (
                f"{macro_name}: CPRS - CH!B5 deviates from expected counterparty transformation"
            )
            assert actual_vols == expected_vols, (
                f"{macro_name}: CPRS - CH!D10:D20 deviates from expected annualized-volatility "
                "range transformation"
            )
            assert actual_allocations == expected_allocations, (
                f"{macro_name}: CPRS - CH!E10:E20 deviates from expected notional-allocation "
                "range transformation"
            )
        finally:
            workbook.close()


@pytest.mark.parametrize("column_letter", ["D", "E"])
def test_macro_spec_cell_by_cell_transformations_match_expected_values(
    macro_spec_cases: tuple[Any, ...],
    parsed_macro_spec_data: dict[str, Any],
    column_letter: str,
) -> None:
    for case in macro_spec_cases:
        parsed = parsed_macro_spec_data[case.variant]
        workbook = case.generator(case.raw_input_path)
        try:
            worksheet = workbook["CPRS - CH"]
            macro_name = _MACRO_BY_VARIANT[case.variant]

            if column_letter == "D":
                expected_values = _pad_to_slot_count(
                    [row.annualized_volatility for row in parsed.totals_rows], 11
                )
                transform_name = "annualized volatility"
            else:
                expected_values = _pad_to_slot_count(_expected_allocations(parsed), 11)
                transform_name = "allocation percentage"

            for offset, expected in enumerate(expected_values):
                row_number = 10 + offset
                cell_ref = f"{column_letter}{row_number}"
                actual = worksheet[cell_ref].value
                assert _values_match(actual=actual, expected=expected), (
                    f"{macro_name}: {cell_ref} deviates from expected {transform_name} "
                    f"transformation (expected={expected!r}, actual={actual!r})"
                )
        finally:
            workbook.close()


def test_macro_spec_invariant_no_blank_numeric_cells_in_core_ranges(
    macro_spec_cases: tuple[Any, ...],
) -> None:
    for case in macro_spec_cases:
        workbook = case.generator(case.raw_input_path)
        try:
            worksheet = workbook["CPRS - CH"]
            macro_name = _MACRO_BY_VARIANT[case.variant]

            for column_letter in ("D", "E"):
                for row_number in range(10, 21):
                    cell_ref = f"{column_letter}{row_number}"
                    value = worksheet[cell_ref].value
                    assert value not in (None, ""), (
                        f"{macro_name}: invariant violated; expected numeric value in "
                        f"CPRS - CH!{cell_ref}, found blank cell"
                    )
        finally:
            workbook.close()


def test_macro_spec_invariant_headers_match_expected_template(
    macro_spec_cases: tuple[Any, ...],
) -> None:
    from openpyxl import load_workbook  # type: ignore[import-untyped]

    template_workbook = load_workbook(_TEMPLATE_WORKBOOK_PATH, data_only=True)
    try:
        template_headers = _read_column_values(template_workbook["CPRS - CH"], "C", 10, 20)
    finally:
        template_workbook.close()

    for case in macro_spec_cases:
        workbook = case.generator(case.raw_input_path)
        try:
            worksheet = workbook["CPRS - CH"]
            macro_name = _MACRO_BY_VARIANT[case.variant]
            output_headers = _read_column_values(worksheet, "C", 10, 20)

            assert output_headers == template_headers, (
                f"{macro_name}: invariant violated; CPRS - CH!C10:C20 headers no longer match "
                "expected MOSERS template headers"
            )
        finally:
            workbook.close()


def _expected_allocations(parsed_data: Any) -> list[float]:
    total_notional = sum(row.notional for row in parsed_data.totals_rows)
    if total_notional == 0:
        return [0.0 for _ in parsed_data.totals_rows]
    return [row.notional / total_notional for row in parsed_data.totals_rows]


def _pad_to_slot_count(values: list[float], slots: int) -> list[float | None]:
    if len(values) >= slots:
        return [*values[:slots]]
    return [*values, *([None] * (slots - len(values)))]


def _read_column_values(
    worksheet: Any, column: str, start_row: int, end_row: int
) -> list[float | str | None]:
    return [
        worksheet[f"{column}{row_number}"].value for row_number in range(start_row, end_row + 1)
    ]


def _values_match(*, actual: Any, expected: Any) -> bool:
    if expected is None:
        return bool(actual in (None, ""))
    if isinstance(expected, float):
        if actual in (None, ""):
            return False
        return math.isclose(float(actual), expected, rel_tol=1e-12, abs_tol=1e-12)
    return bool(actual == expected)
