"""Workbook comparison helpers for strict range-level assertions in tests."""

from __future__ import annotations

from collections.abc import Mapping, Sequence
from pathlib import Path
from typing import Any

from tests.helpers.excel_asserts import parse_a1_range

WorkbookSource = Any


def compare_workbooks(
    *,
    reference_workbook: WorkbookSource,
    generated_workbook: WorkbookSource,
    sheet_names: Sequence[str],
    ranges_by_sheet: Mapping[str, str | Sequence[str]],
) -> None:
    """Compare workbook values and style attributes for specific sheet ranges."""

    reference, close_reference = _coerce_workbook(reference_workbook)
    generated, close_generated = _coerce_workbook(generated_workbook)
    try:
        for sheet_name in sheet_names:
            if sheet_name not in reference.sheetnames:
                raise AssertionError(f"Missing sheet in reference workbook: {sheet_name}")
            if sheet_name not in generated.sheetnames:
                raise AssertionError(f"Missing sheet in generated workbook: {sheet_name}")

            reference_sheet = reference[sheet_name]
            generated_sheet = generated[sheet_name]
            for range_ref in _coerce_ranges(ranges_by_sheet, sheet_name):
                bounds = parse_a1_range(range_ref)
                for row_number in range(bounds.min_row, bounds.max_row + 1):
                    for column_number in range(bounds.min_col, bounds.max_col + 1):
                        reference_cell = reference_sheet.cell(row=row_number, column=column_number)
                        generated_cell = generated_sheet.cell(row=row_number, column=column_number)
                        coordinate = reference_cell.coordinate

                        _assert_cell_property(
                            sheet_name,
                            coordinate,
                            "value",
                            reference_cell.value,
                            generated_cell.value,
                        )
                        _assert_cell_property(
                            sheet_name,
                            coordinate,
                            "number_format",
                            reference_cell.number_format,
                            generated_cell.number_format,
                        )
                        _assert_cell_property(
                            sheet_name,
                            coordinate,
                            "font",
                            _normalize_font(reference_cell.font),
                            _normalize_font(generated_cell.font),
                        )
                        _assert_cell_property(
                            sheet_name,
                            coordinate,
                            "fill",
                            _normalize_fill(reference_cell.fill),
                            _normalize_fill(generated_cell.fill),
                        )
                        _assert_cell_property(
                            sheet_name,
                            coordinate,
                            "alignment",
                            _normalize_alignment(reference_cell.alignment),
                            _normalize_alignment(generated_cell.alignment),
                        )
                        _assert_cell_property(
                            sheet_name,
                            coordinate,
                            "border",
                            _normalize_border(reference_cell.border),
                            _normalize_border(generated_cell.border),
                        )
    finally:
        if close_reference:
            reference.close()
        if close_generated:
            generated.close()


def _coerce_workbook(workbook_source: WorkbookSource) -> tuple[Any, bool]:
    if hasattr(workbook_source, "sheetnames") and hasattr(workbook_source, "close"):
        return workbook_source, False

    path = Path(workbook_source)
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ModuleNotFoundError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required for workbook comparison") from exc

    workbook = load_workbook(filename=path, read_only=False, data_only=False)
    return workbook, True


def _coerce_ranges(
    ranges_by_sheet: Mapping[str, str | Sequence[str]], sheet_name: str
) -> tuple[str, ...]:
    if sheet_name not in ranges_by_sheet:
        raise AssertionError(f"No ranges defined for sheet: {sheet_name}")

    range_spec = ranges_by_sheet[sheet_name]
    if isinstance(range_spec, str):
        return (range_spec,)
    return tuple(range_spec)


def _assert_cell_property(
    sheet_name: str, coordinate: str, property_name: str, expected: Any, actual: Any
) -> None:
    if expected != actual:
        raise AssertionError(
            "Cell mismatch "
            f"sheet='{sheet_name}' "
            f"cell='{coordinate}' "
            f"property='{property_name}' "
            f"expected={expected!r} "
            f"actual={actual!r}"
        )


def _normalize_font(font: Any) -> tuple[Any, ...]:
    return (
        getattr(font, "name", None),
        getattr(font, "size", None),
        getattr(font, "bold", None),
        getattr(font, "italic", None),
        getattr(font, "underline", None),
        getattr(font, "strike", None),
        getattr(font, "vertAlign", None),
        getattr(font, "outline", None),
        getattr(font, "shadow", None),
        getattr(font, "charset", None),
        getattr(font, "family", None),
        getattr(font, "scheme", None),
        _normalize_color(getattr(font, "color", None)),
    )


def _normalize_fill(fill: Any) -> tuple[Any, ...]:
    return (
        getattr(fill, "fill_type", None),
        getattr(fill, "patternType", None),
        _normalize_color(getattr(fill, "fgColor", None)),
        _normalize_color(getattr(fill, "bgColor", None)),
    )


def _normalize_alignment(alignment: Any) -> tuple[Any, ...]:
    return (
        getattr(alignment, "horizontal", None),
        getattr(alignment, "vertical", None),
        getattr(alignment, "text_rotation", None),
        getattr(alignment, "wrap_text", None),
        getattr(alignment, "shrink_to_fit", None),
        getattr(alignment, "indent", None),
        getattr(alignment, "relativeIndent", None),
        getattr(alignment, "justifyLastLine", None),
        getattr(alignment, "readingOrder", None),
    )


def _normalize_border(border: Any) -> tuple[Any, ...]:
    return (
        _normalize_side(getattr(border, "left", None)),
        _normalize_side(getattr(border, "right", None)),
        _normalize_side(getattr(border, "top", None)),
        _normalize_side(getattr(border, "bottom", None)),
        _normalize_side(getattr(border, "diagonal", None)),
        _normalize_side(getattr(border, "vertical", None)),
        _normalize_side(getattr(border, "horizontal", None)),
        getattr(border, "diagonalDown", None),
        getattr(border, "diagonalUp", None),
        getattr(border, "outline", None),
    )


def _normalize_side(side: Any) -> tuple[Any, ...]:
    if side is None:
        return (None, None)
    return (getattr(side, "style", None), _normalize_color(getattr(side, "color", None)))


def _normalize_color(color: Any) -> tuple[Any, ...]:
    if color is None:
        return (None, None, None, None)
    return (
        getattr(color, "type", None),
        getattr(color, "rgb", None),
        getattr(color, "indexed", None),
        getattr(color, "tint", None),
    )
