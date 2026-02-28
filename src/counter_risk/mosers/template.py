"""Centralized retrieval helpers for the internal MOSERS template workbook."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

_TEMPLATE_NAME = "mosers_template.xlsx"


def get_mosers_template_path() -> Path:
    """Return the filesystem path to the bundled MOSERS workbook template."""

    template_path = Path(__file__).with_name(_TEMPLATE_NAME)
    if not template_path.exists():
        raise FileNotFoundError(f"MOSERS template workbook not found: {template_path}")
    return template_path


def get_mosers_template_bytes() -> bytes:
    """Return the raw bytes for the bundled MOSERS workbook template."""

    return get_mosers_template_path().read_bytes()


def load_mosers_template_workbook() -> Any:
    """Load the internal MOSERS template workbook into an editable openpyxl workbook."""

    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ModuleNotFoundError as exc:  # pragma: no cover - environment dependent
        raise RuntimeError("openpyxl is required to load MOSERS template workbooks") from exc

    return load_workbook(filename=get_mosers_template_path())


@dataclass(frozen=True)
class MarkerBoundRange:
    """Inclusive row range discovered from text markers."""

    start_row: int
    end_row: int


def normalize_template_text(value: object) -> str:
    """Return normalized text for resilient marker matching."""

    return " ".join(str(value or "").split()).strip().casefold()


def find_row_containing_text(
    worksheet: Any,
    text: str,
    *,
    min_row: int = 1,
    max_row: int | None = None,
) -> int | None:
    """Return the first row in *worksheet* containing *text* in any column."""

    if min_row < 1:
        min_row = 1
    end_row = int(worksheet.max_row) if max_row is None else max(1, int(max_row))
    if end_row < min_row:
        return None

    marker = normalize_template_text(text)
    if not marker:
        return None

    max_column = int(worksheet.max_column)
    for row_number in range(min_row, end_row + 1):
        for column_number in range(1, max_column + 1):
            value = worksheet.cell(row=row_number, column=column_number).value
            if marker in normalize_template_text(value):
                return row_number
    return None


def resolve_marker_bound_range(
    worksheet: Any,
    *,
    start_marker: str,
    end_marker: str,
    min_row: int = 1,
    max_row: int | None = None,
) -> MarkerBoundRange:
    """Resolve inclusive row bounds by locating *start_marker* and *end_marker*."""

    start_row = find_row_containing_text(
        worksheet, start_marker, min_row=min_row, max_row=max_row
    )
    if start_row is None:
        raise ValueError(
            f"Unable to locate marker {start_marker!r} in sheet {worksheet.title!r}."
        )
    end_row = find_row_containing_text(
        worksheet,
        end_marker,
        min_row=start_row,
        max_row=max_row,
    )
    if end_row is None:
        raise ValueError(f"Unable to locate marker {end_marker!r} in sheet {worksheet.title!r}.")
    if end_row < start_row:
        raise ValueError(
            f"Marker ordering invalid in sheet {worksheet.title!r}: "
            f"start marker {start_marker!r} occurs after end marker {end_marker!r}."
        )
    return MarkerBoundRange(start_row=start_row, end_row=end_row)
