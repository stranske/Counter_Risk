"""Parser for raw NISA "Monthly All Programs" workbooks."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

_TOTALS_MARKER = "total by counterparty/clearing house"
_TOTALS_STOP_MARKERS = (
    "total current exposure",
    "mosers program",
    "notional breakdown",
)

_SEGMENT_MAP = {
    "swaps": "swaps",
    "repo": "repo",
    "futures / cdx": "futures_cdx",
    "futures/cdx": "futures_cdx",
    "futures cdx": "futures_cdx",
    "futures": "futures",
}


@dataclass(frozen=True)
class NisaChRow:
    """Normalized row from the top CPRS-CH segment block."""

    segment: str
    counterparty: str
    cash: float
    tips: float
    treasury: float
    equity: float
    commodity: float
    currency: float
    notional: float
    notional_change: float
    annualized_volatility: float


@dataclass(frozen=True)
class NisaTotalsRow:
    """Normalized row from totals-by-counterparty block."""

    counterparty: str
    tips: float
    treasury: float
    equity: float
    commodity: float
    currency: float
    notional: float
    notional_change: float
    annualized_volatility: float


@dataclass(frozen=True)
class NisaAllProgramsData:
    """Parsed payload for one raw NISA Monthly All Programs workbook."""

    ch_rows: tuple[NisaChRow, ...]
    totals_rows: tuple[NisaTotalsRow, ...]


def parse_nisa_all_programs(path: Path | str) -> NisaAllProgramsData:
    """Parse raw NISA workbook into deterministic CH and totals row sets."""

    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"NISA raw workbook not found: {workbook_path}")
    if workbook_path.suffix.lower() != ".xlsx":
        raise ValueError(f"NISA raw workbook must be an .xlsx file: {workbook_path}")

    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ModuleNotFoundError as exc:  # pragma: no cover - environment dependent
        raise RuntimeError("openpyxl is required to parse NISA workbooks") from exc

    try:
        workbook = load_workbook(filename=workbook_path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Unable to open NISA workbook: {workbook_path}") from exc

    try:
        if not workbook.sheetnames:
            raise ValueError("NISA workbook contains no worksheets")
        worksheet = workbook[workbook.sheetnames[0]]
        totals_marker_row = _find_totals_marker_row(worksheet=worksheet)
        if totals_marker_row is None:
            raise ValueError("Unable to locate 'Total by Counterparty/Clearing House' section")
        ch_rows = _parse_ch_rows(worksheet=worksheet, totals_marker_row=totals_marker_row)
        totals_rows = _parse_totals_rows(worksheet=worksheet, totals_marker_row=totals_marker_row)
    finally:
        workbook.close()

    if not ch_rows:
        raise ValueError("NISA workbook parser produced no CPRS-CH rows")
    if not totals_rows:
        raise ValueError("NISA workbook parser produced no totals rows")

    return NisaAllProgramsData(ch_rows=tuple(ch_rows), totals_rows=tuple(totals_rows))


def _normalize_text(value: Any) -> str:
    return " ".join(str(value or "").split()).strip()


def _coerce_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = _normalize_text(value)
    if not text or text in {"-", "--", "N/A", "n/a"}:
        return 0.0
    if text.startswith("(") and text.endswith(")"):
        text = f"-{text[1:-1]}"

    cleaned = text.replace(",", "").replace("$", "").replace("%", "")
    try:
        return float(cleaned)
    except ValueError as exc:
        raise ValueError(f"Unable to parse numeric cell value: {value!r}") from exc


def _find_totals_marker_row(*, worksheet: Any) -> int | None:
    for row_number in range(1, int(worksheet.max_row) + 1):
        marker_text = _normalize_text(worksheet.cell(row=row_number, column=2).value).casefold()
        if _TOTALS_MARKER in marker_text:
            return row_number
    return None


def _parse_ch_rows(*, worksheet: Any, totals_marker_row: int) -> list[NisaChRow]:
    rows: list[NisaChRow] = []
    current_segment = ""

    for row_number in range(1, totals_marker_row):
        segment_text = _normalize_text(worksheet.cell(row=row_number, column=1).value)
        mapped_segment = _SEGMENT_MAP.get(segment_text.casefold())
        if mapped_segment is not None:
            current_segment = mapped_segment

        counterparty = _normalize_text(worksheet.cell(row=row_number, column=2).value)
        if not counterparty:
            continue
        if counterparty.casefold() in {"total", "subtotal"}:
            continue
        if not current_segment:
            continue

        rows.append(
            NisaChRow(
                segment=current_segment,
                counterparty=counterparty,
                cash=_coerce_float(worksheet.cell(row=row_number, column=4).value),
                tips=_coerce_float(worksheet.cell(row=row_number, column=5).value),
                treasury=_coerce_float(worksheet.cell(row=row_number, column=6).value),
                equity=_coerce_float(worksheet.cell(row=row_number, column=7).value),
                commodity=_coerce_float(worksheet.cell(row=row_number, column=8).value),
                currency=_coerce_float(worksheet.cell(row=row_number, column=9).value),
                notional=_coerce_float(worksheet.cell(row=row_number, column=11).value),
                notional_change=_coerce_float(worksheet.cell(row=row_number, column=12).value),
                annualized_volatility=_coerce_float(
                    worksheet.cell(row=row_number, column=14).value
                ),
            )
        )

    return rows


def _parse_totals_rows(*, worksheet: Any, totals_marker_row: int) -> list[NisaTotalsRow]:
    rows: list[NisaTotalsRow] = []

    for row_number in range(totals_marker_row + 1, int(worksheet.max_row) + 1):
        row_label = _normalize_text(worksheet.cell(row=row_number, column=2).value).casefold()
        if any(marker in row_label for marker in _TOTALS_STOP_MARKERS):
            break

        counterparty = _normalize_text(worksheet.cell(row=row_number, column=2).value)
        if not counterparty:
            continue
        if counterparty.casefold() in {"total", "subtotal"}:
            continue

        rows.append(
            NisaTotalsRow(
                counterparty=counterparty,
                tips=_coerce_float(worksheet.cell(row=row_number, column=5).value),
                treasury=_coerce_float(worksheet.cell(row=row_number, column=6).value),
                equity=_coerce_float(worksheet.cell(row=row_number, column=7).value),
                commodity=_coerce_float(worksheet.cell(row=row_number, column=8).value),
                currency=_coerce_float(worksheet.cell(row=row_number, column=9).value),
                notional=_coerce_float(worksheet.cell(row=row_number, column=11).value),
                notional_change=_coerce_float(worksheet.cell(row=row_number, column=12).value),
                annualized_volatility=_coerce_float(
                    worksheet.cell(row=row_number, column=14).value
                ),
            )
        )

    return rows
