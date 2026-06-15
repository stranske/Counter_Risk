"""Structured Repo Cash source and overrides loaders."""

from __future__ import annotations

import csv
from collections.abc import Callable
from pathlib import Path
from typing import Literal

from counter_risk.normalize import normalize_counterparty

_COUNTERPARTY_HEADER_ALIASES = ("counterparty", "counterparty_name", "name")
_CASH_HEADER_ALIASES = ("cash_value", "cash", "repo_cash")
_OVERRIDES_REQUIRED_HEADERS = ("counterparty", "cash_value")


def load_repo_cash_structured_source(
    path: Path | str,
    *,
    source_type: Literal["csv", "xlsx"] | None = None,
) -> dict[str, float]:
    source_path = Path(path)
    resolved_source_type = _resolve_source_type(source_path, source_type)
    if resolved_source_type == "csv":
        rows = _load_csv_rows(source_path)
    else:
        rows = _load_xlsx_rows(source_path)
    return _rows_to_repo_cash_mapping(rows, path=source_path)


def load_repo_cash_overrides_csv(path: Path | str) -> tuple[dict[str, float], list[dict[str, str]]]:
    overrides_path = Path(path)
    rows = _load_csv_rows(overrides_path)
    if not rows:
        return {}, []

    first_row = rows[0]
    normalized_headers = {key.strip().casefold() for key in first_row}
    missing_headers = [key for key in _OVERRIDES_REQUIRED_HEADERS if key not in normalized_headers]
    if missing_headers:
        raise ValueError(
            f"Overrides file '{overrides_path}' is missing required columns: {', '.join(missing_headers)}"
        )

    overrides: dict[str, float] = {}
    audit_rows: list[dict[str, str]] = []
    for row_index, raw_row in enumerate(rows, start=2):
        row = {_normalize_header(key): value for key, value in raw_row.items()}
        raw_counterparty = row.get("counterparty", "").strip()
        if not raw_counterparty:
            continue
        raw_cash_value = row.get("cash_value", "").strip()
        if not raw_cash_value:
            continue
        cash_value = _coerce_cash_value(raw_cash_value, path=overrides_path, row_index=row_index)
        canonical_counterparty = normalize_counterparty(raw_counterparty)
        overrides[canonical_counterparty] = cash_value
        audit_rows.append(
            {
                "counterparty": canonical_counterparty,
                "raw_counterparty": raw_counterparty,
                "cash_value": f"{cash_value}",
                "note": row.get("note", "").strip(),
            }
        )
    return overrides, audit_rows


def load_cash_by_counterparty(
    *,
    source_type: Literal["csv", "xlsx", "pdf", "none"] | None,
    source_path: Path | str | None,
    overrides_path: Path | str | None = None,
    pdf_parser: Callable[[Path], dict[str, float]] | None = None,
) -> tuple[dict[str, float], list[dict[str, str]], str, list[str]]:
    """Load repo cash indexed by canonical counterparty name.

    Precedence (highest to lowest):
    - Override file values are applied per-counterparty on top of the base source.
    - Structured source (CSV/XLSX) or PDF parser provides the base values.

    Returns a 4-tuple:
      cash_by_counterparty   – final mapping after overrides are applied
      override_audit_rows    – audit records for each applied override row
      source_label           – description of the base source used (e.g. ``"csv:/path/file.csv"``)
      duplicate_counterparty_names – canonical names that appeared >1 time in the base source
    """
    resolved_type = (source_type or "").strip().casefold()
    resolved_source = Path(source_path) if source_path is not None else None

    if resolved_type == "none" or resolved_source is None:
        return {}, [], "none", []

    effective_type = resolved_type
    if effective_type not in {"pdf", "csv", "xlsx"}:
        suffix = resolved_source.suffix.lower()
        if suffix == ".csv":
            effective_type = "csv"
        elif suffix in {".xlsx", ".xlsm"}:
            effective_type = "xlsx"
        elif suffix == ".pdf":
            effective_type = "pdf"
        else:
            raise ValueError(f"Cannot determine cash source type for '{resolved_source}'.")

    duplicate_names: list[str] = []
    if effective_type == "pdf":
        if pdf_parser is None:
            from counter_risk.parsers.daily_holdings_pdf import parse_daily_holdings_pdf

            pdf_parser = parse_daily_holdings_pdf
        repo_cash: dict[str, float] = pdf_parser(resolved_source)
    else:
        rows = (
            _load_csv_rows(resolved_source)
            if effective_type == "csv"
            else _load_xlsx_rows(resolved_source)
        )
        duplicate_names = _find_duplicate_counterparty_names_in_rows(rows)
        repo_cash = _rows_to_repo_cash_mapping(rows, path=resolved_source)

    source_label = f"{effective_type}:{resolved_source}"
    audit_rows: list[dict[str, str]] = []

    if overrides_path is not None:
        resolved_overrides = Path(overrides_path)
        overrides, override_audit_rows = load_repo_cash_overrides_csv(resolved_overrides)
        if overrides:
            repo_cash.update(overrides)
            audit_rows = override_audit_rows

    return repo_cash, audit_rows, source_label, duplicate_names


def find_duplicate_counterparty_names(
    path: Path | str,
    *,
    source_type: Literal["csv", "xlsx"] | None = None,
) -> list[str]:
    """Return canonical counterparty names that appear more than once in a structured source file."""
    source_path = Path(path)
    resolved_type = _resolve_source_type(source_path, source_type)
    rows = _load_csv_rows(source_path) if resolved_type == "csv" else _load_xlsx_rows(source_path)
    return _find_duplicate_counterparty_names_in_rows(rows)


def _find_duplicate_counterparty_names_in_rows(rows: list[dict[str, str]]) -> list[str]:
    normalized_rows = [{_normalize_header(k): v for k, v in row.items()} for row in rows]
    header_names = {k for row in normalized_rows for k in row}
    counterparty_header = _first_matching_header(header_names, _COUNTERPARTY_HEADER_ALIASES)
    if counterparty_header is None:
        return []
    seen: dict[str, int] = {}
    for row in normalized_rows:
        raw = str(row.get(counterparty_header, "")).strip()
        if not raw:
            continue
        canonical = normalize_counterparty(raw)
        seen[canonical] = seen.get(canonical, 0) + 1
    return sorted(name for name, count in seen.items() if count > 1)


def _resolve_source_type(
    source_path: Path,
    source_type: Literal["csv", "xlsx"] | None,
) -> Literal["csv", "xlsx"]:
    suffix = source_path.suffix.lower()
    if source_type is not None:
        if source_type == "csv" and suffix not in {"", ".csv"}:
            raise ValueError(
                f"cash_source_type=csv requires a .csv file path, got '{source_path}'."
            )
        if source_type == "xlsx" and suffix not in {"", ".xlsx", ".xlsm"}:
            raise ValueError(
                f"cash_source_type=xlsx requires a .xlsx/.xlsm file path, got '{source_path}'."
            )
        return source_type
    if suffix == ".csv":
        return "csv"
    if suffix in {".xlsx", ".xlsm"}:
        return "xlsx"
    raise ValueError(f"Unsupported cash source extension for '{source_path}'.")


def _load_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows: list[dict[str, str]] = []
        for row in reader:
            if row is None:
                continue
            normalized_row = {
                key: "" if value is None else str(value)
                for key, value in row.items()
                if isinstance(key, str)
            }
            if normalized_row:
                rows.append(normalized_row)
        return rows


def _load_xlsx_rows(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:  # pragma: no cover - exercised in dependency-absence tests
        raise RuntimeError(
            "openpyxl is required to load structured Repo Cash XLSX sources."
        ) from exc

    workbook = load_workbook(filename=path, data_only=True, read_only=True)
    try:
        worksheet = workbook.active
        iterator = worksheet.iter_rows(values_only=True)
        try:
            raw_headers = next(iterator)
        except StopIteration:
            return []
        headers = ["" if value is None else str(value).strip() for value in raw_headers]
        rows: list[dict[str, str]] = []
        for raw_row in iterator:
            row: dict[str, str] = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                cell_value = raw_row[index] if index < len(raw_row) else None
                row[header] = "" if cell_value is None else str(cell_value).strip()
            if row:
                rows.append(row)
        return rows
    finally:
        workbook.close()


def _rows_to_repo_cash_mapping(rows: list[dict[str, str]], *, path: Path) -> dict[str, float]:
    if not rows:
        return {}
    normalized_rows = [
        {_normalize_header(key): value for key, value in row.items()} for row in rows
    ]
    header_names = {key for row in normalized_rows for key in row}
    counterparty_header = _first_matching_header(header_names, _COUNTERPARTY_HEADER_ALIASES)
    cash_header = _first_matching_header(header_names, _CASH_HEADER_ALIASES)
    if counterparty_header is None or cash_header is None:
        raise ValueError(
            f"Structured cash source '{path}' must include counterparty and cash columns."
        )

    mapping: dict[str, float] = {}
    for row_index, row in enumerate(normalized_rows, start=2):
        raw_counterparty = str(row.get(counterparty_header, "")).strip()
        if not raw_counterparty:
            continue
        raw_cash = str(row.get(cash_header, "")).strip()
        if not raw_cash:
            continue
        canonical_counterparty = normalize_counterparty(raw_counterparty)
        mapping[canonical_counterparty] = _coerce_cash_value(
            raw_cash, path=path, row_index=row_index
        )
    return mapping


def _normalize_header(value: str) -> str:
    return value.strip().casefold().replace(" ", "_")


def _first_matching_header(headers: set[str], candidates: tuple[str, ...]) -> str | None:
    for candidate in candidates:
        if candidate in headers:
            return candidate
    return None


def _coerce_cash_value(raw_value: str, *, path: Path, row_index: int) -> float:
    normalized = raw_value.replace(",", "").strip()
    try:
        return float(normalized)
    except ValueError as exc:
        raise ValueError(
            f"Invalid cash value '{raw_value}' in '{path}' at row {row_index}."
        ) from exc
