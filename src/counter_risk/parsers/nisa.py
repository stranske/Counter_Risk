"""Parser for raw NISA "Monthly All Programs" workbooks."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

_TOTALS_MARKER = "total by counterparty/clearing house"
_TOTALS_STOP_MARKERS = (
    "total current exposure",
    "mosers program",
    "notional breakdown",
)

_SEGMENT_MAP = {
    "swaps": "swaps",
    "repo": "repo",
    "futures / cdx": "futures_cdx",
    "futures/cdx": "futures_cdx",
    "futures cdx": "futures_cdx",
    "futures": "futures",
}

_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "counterparty": (
        "counterparty/clearing house",
        "counterparty/ clearing house",
        "counterparty / clearing house",
        "counterparty /clearing house",
        "counterparty clearing house",
        "counterparty/fcm",
        "counterparty/ fcm",
    ),
    "cash": ("cash",),
    "tips": ("tips",),
    "treasury": ("treasury",),
    "equity": ("equity",),
    "commodity": ("commodity",),
    "currency": ("currency",),
    "notional": ("notional",),
    "notional_change": (
        "notional change from prior month",
        "notional change",
        "from prior month***",
        "from prior month",
    ),
    "annualized_volatility": ("annualized volatility",),
}

_REQUIRED_HEADERS: tuple[str, ...] = (
    "cash",
    "tips",
    "treasury",
    "equity",
    "commodity",
    "currency",
    "notional",
    "notional_change",
    "annualized_volatility",
)

_HEADER_SCAN_ROW_LIMIT = 200


@dataclass(frozen=True)
class NisaChRow:
    """Normalized row from the top CPRS-CH segment block."""

    segment: str
    counterparty: str
    cash: float
    tips: float
    treasury: float
    equity: float
    commodity: float
    currency: float
    notional: float
    notional_change: float
    annualized_volatility: float


@dataclass(frozen=True)
class NisaTotalsRow:
    """Normalized row from totals-by-counterparty block."""

    counterparty: str
    tips: float
    treasury: float
    equity: float
    commodity: float
    currency: float
    notional: float
    notional_change: float
    annualized_volatility: float


@dataclass(frozen=True)
class NisaAllProgramsData:
    """Parsed payload for one raw NISA Monthly All Programs workbook."""

    ch_rows: tuple[NisaChRow, ...]
    totals_rows: tuple[NisaTotalsRow, ...]


def parse_nisa_all_programs(path: Path | str) -> NisaAllProgramsData:
    """Parse raw NISA workbook into deterministic CH and totals row sets."""

    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"NISA raw workbook not found: {workbook_path}")
    if workbook_path.suffix.lower() != ".xlsx":
        raise ValueError(f"NISA raw workbook must be an .xlsx file: {workbook_path}")

    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ModuleNotFoundError as exc:  # pragma: no cover - environment dependent
        raise RuntimeError("openpyxl is required to parse NISA workbooks") from exc

    try:
        workbook = load_workbook(filename=workbook_path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Unable to open NISA workbook: {workbook_path}") from exc

    try:
        if not workbook.sheetnames:
            raise ValueError("NISA workbook contains no worksheets")
        worksheet, header_row, header_columns = _select_worksheet_and_headers(workbook=workbook)
        totals_marker_row = _find_totals_marker_row(
            worksheet=worksheet, counterparty_column=header_columns["counterparty"]
        )
        if totals_marker_row is None:
            raise ValueError("Unable to locate 'Total by Counterparty/Clearing House' section")
        ch_rows = _parse_ch_rows(
            worksheet=worksheet,
            totals_marker_row=totals_marker_row,
            header_row=header_row,
            header_columns=header_columns,
        )
        totals_rows = _parse_totals_rows(
            worksheet=worksheet,
            totals_marker_row=totals_marker_row,
            header_columns=header_columns,
        )
    finally:
        workbook.close()

    if not ch_rows:
        raise ValueError("NISA workbook parser produced no CPRS-CH rows")
    if not totals_rows:
        raise ValueError("NISA workbook parser produced no totals rows")

    return NisaAllProgramsData(ch_rows=tuple(ch_rows), totals_rows=tuple(totals_rows))


def _normalize_text(value: Any) -> str:
    return " ".join(str(value or "").split()).strip()


def _coerce_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = _normalize_text(value)
    if not text or text in {"-", "--", "N/A", "n/a"}:
        return 0.0
    if text.startswith("(") and text.endswith(")"):
        text = f"-{text[1:-1]}"

    cleaned = text.replace(",", "").replace("$", "").replace("%", "")
    try:
        return float(cleaned)
    except ValueError as exc:
        raise ValueError(f"Unable to parse numeric cell value: {value!r}") from exc


def _find_totals_marker_row(*, worksheet: Any, counterparty_column: int) -> int | None:
    for row_number in range(1, int(worksheet.max_row) + 1):
        marker_text = _normalize_text(
            worksheet.cell(row=row_number, column=counterparty_column).value
        ).casefold()
        if _TOTALS_MARKER in marker_text:
            return row_number
    return None


def _select_worksheet_and_headers(*, workbook: Any) -> tuple[Any, int, dict[str, int]]:
    candidates: list[tuple[int, int, str, Any, dict[str, int], tuple[str, ...]]] = []
    evaluated_sheet_names: list[str] = []

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        header_row, header_columns, missing = _find_header_row_and_columns(worksheet=worksheet)
        score = len(_REQUIRED_HEADERS) - len(missing)
        candidates.append((score, header_row, sheet_name, worksheet, header_columns, missing))

    candidates.sort(key=lambda item: (-item[0], item[1], item[2].casefold()))

    for _score, header_row, sheet_name, worksheet, header_columns, missing in candidates:
        evaluated_sheet_names.append(sheet_name)
        if missing:
            continue

        counterparty_column = _infer_counterparty_column(
            worksheet=worksheet,
            header_columns=header_columns,
        )
        if counterparty_column is None:
            continue

        resolved_columns = dict(header_columns)
        resolved_columns["counterparty"] = counterparty_column
        return worksheet, header_row, resolved_columns

    required_text = ", ".join(_REQUIRED_HEADERS)
    evaluated_text = ", ".join(evaluated_sheet_names) if evaluated_sheet_names else "<none>"
    raise ValueError(
        f"Missing required headers: {required_text}; evaluated sheets: {evaluated_text}"
    )


def _find_header_row_and_columns(*, worksheet: Any) -> tuple[int, dict[str, int], tuple[str, ...]]:
    best_row = 1
    best_columns: dict[str, int] = {}
    best_missing: tuple[str, ...] = _REQUIRED_HEADERS
    max_row = min(int(worksheet.max_row), _HEADER_SCAN_ROW_LIMIT)
    max_column = int(worksheet.max_column)

    for row_number in range(1, max_row + 1):
        columns = _build_header_column_map(
            worksheet=worksheet, header_row=row_number, max_column=max_column
        )
        missing = tuple(header for header in _REQUIRED_HEADERS if header not in columns)
        if len(missing) < len(best_missing):
            best_row = row_number
            best_columns = columns
            best_missing = missing
            if not best_missing:
                break

    return best_row, best_columns, best_missing


def _build_header_column_map(*, worksheet: Any, header_row: int, max_column: int) -> dict[str, int]:
    columns: dict[str, int] = {}
    candidate_rows = [header_row]
    if header_row > 1:
        candidate_rows.append(header_row - 1)
    if header_row < int(worksheet.max_row):
        candidate_rows.append(header_row + 1)

    for column in range(1, max_column + 1):
        pieces = [
            _normalize_text(worksheet.cell(row=row_number, column=column).value).casefold()
            for row_number in candidate_rows
        ]
        column_text = " ".join(piece for piece in pieces if piece)
        if not column_text:
            continue
        for header_name, aliases in _HEADER_ALIASES.items():
            if header_name in columns:
                continue
            if any(alias in column_text for alias in aliases):
                columns[header_name] = column
                break
    return columns


def _infer_counterparty_column(*, worksheet: Any, header_columns: dict[str, int]) -> int | None:
    explicit = header_columns.get("counterparty")
    if explicit is not None:
        return explicit

    marker = _find_totals_marker_position(worksheet=worksheet)
    if marker is not None:
        _row, column = marker
        return column

    if header_columns:
        return max(1, min(header_columns.values()) - 1)
    return None


def _find_totals_marker_position(*, worksheet: Any) -> tuple[int, int] | None:
    max_row = int(worksheet.max_row)
    max_column = int(worksheet.max_column)
    for row_number in range(1, max_row + 1):
        for column_number in range(1, max_column + 1):
            cell_text = _normalize_text(
                worksheet.cell(row=row_number, column=column_number).value
            ).casefold()
            if _TOTALS_MARKER in cell_text:
                return row_number, column_number
    return None


def _parse_ch_rows(
    *,
    worksheet: Any,
    totals_marker_row: int,
    header_row: int,
    header_columns: dict[str, int],
) -> list[NisaChRow]:
    rows: list[NisaChRow] = []
    current_segment = ""
    segment_column = _detect_segment_column(
        worksheet=worksheet,
        header_row=header_row,
        totals_marker_row=totals_marker_row,
        counterparty_column=header_columns["counterparty"],
    )

    for row_number in range(header_row + 1, totals_marker_row):
        segment_text = _normalize_text(worksheet.cell(row=row_number, column=segment_column).value)
        mapped_segment = _SEGMENT_MAP.get(segment_text.casefold())
        if mapped_segment is not None:
            current_segment = mapped_segment

        counterparty = _normalize_text(
            worksheet.cell(row=row_number, column=header_columns["counterparty"]).value
        )
        if not counterparty:
            continue
        if counterparty.casefold() in {"total", "subtotal"}:
            continue
        if not current_segment:
            continue

        rows.append(
            NisaChRow(
                segment=current_segment,
                counterparty=counterparty,
                cash=_coerce_float(
                    worksheet.cell(row=row_number, column=header_columns["cash"]).value
                ),
                tips=_coerce_float(
                    worksheet.cell(row=row_number, column=header_columns["tips"]).value
                ),
                treasury=_coerce_float(
                    worksheet.cell(row=row_number, column=header_columns["treasury"]).value
                ),
                equity=_coerce_float(
                    worksheet.cell(row=row_number, column=header_columns["equity"]).value
                ),
                commodity=_coerce_float(
                    worksheet.cell(row=row_number, column=header_columns["commodity"]).value
                ),
                currency=_coerce_float(
                    worksheet.cell(row=row_number, column=header_columns["currency"]).value
                ),
                notional=_coerce_float(
                    worksheet.cell(row=row_number, column=header_columns["notional"]).value
                ),
                notional_change=_coerce_float(
                    worksheet.cell(row=row_number, column=header_columns["notional_change"]).value
                ),
                annualized_volatility=_coerce_float(
                    worksheet.cell(
                        row=row_number, column=header_columns["annualized_volatility"]
                    ).value
                ),
            )
        )

    return rows


def _detect_segment_column(
    *,
    worksheet: Any,
    header_row: int,
    totals_marker_row: int,
    counterparty_column: int,
) -> int:
    candidates = (max(1, counterparty_column - 1), 1)
    best_column = candidates[0]
    best_score = -1
    for column in candidates:
        score = 0
        for row_number in range(header_row + 1, totals_marker_row):
            segment_text = _normalize_text(
                worksheet.cell(row=row_number, column=column).value
            ).casefold()
            if segment_text in _SEGMENT_MAP:
                score += 1
        if score > best_score:
            best_column = column
            best_score = score
    return best_column


def _parse_totals_rows(
    *,
    worksheet: Any,
    totals_marker_row: int,
    header_columns: dict[str, int],
) -> list[NisaTotalsRow]:
    rows: list[NisaTotalsRow] = []

    for row_number in range(totals_marker_row + 1, int(worksheet.max_row) + 1):
        row_label = _normalize_text(
            worksheet.cell(row=row_number, column=header_columns["counterparty"]).value
        ).casefold()
        if any(marker in row_label for marker in _TOTALS_STOP_MARKERS):
            break

        counterparty = _normalize_text(
            worksheet.cell(row=row_number, column=header_columns["counterparty"]).value
        )
        if not counterparty:
            continue
        if counterparty.casefold() in {"total", "subtotal"}:
            continue

        rows.append(
            NisaTotalsRow(
                counterparty=counterparty,
                tips=_coerce_float(
                    worksheet.cell(row=row_number, column=header_columns["tips"]).value
                ),
                treasury=_coerce_float(
                    worksheet.cell(row=row_number, column=header_columns["treasury"]).value
                ),
                equity=_coerce_float(
                    worksheet.cell(row=row_number, column=header_columns["equity"]).value
                ),
                commodity=_coerce_float(
                    worksheet.cell(row=row_number, column=header_columns["commodity"]).value
                ),
                currency=_coerce_float(
                    worksheet.cell(row=row_number, column=header_columns["currency"]).value
                ),
                notional=_coerce_float(
                    worksheet.cell(row=row_number, column=header_columns["notional"]).value
                ),
                notional_change=_coerce_float(
                    worksheet.cell(row=row_number, column=header_columns["notional_change"]).value
                ),
                annualized_volatility=_coerce_float(
                    worksheet.cell(
                        row=row_number, column=header_columns["annualized_volatility"]
                    ).value
                ),
            )
        )

    return rows


__all__ = [
    "NisaAllProgramsData",
    "NisaChRow",
    "NisaTotalsRow",
    "parse_nisa_all_programs",
]
