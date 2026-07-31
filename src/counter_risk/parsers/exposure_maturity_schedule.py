"""Parser for the NISA Monthly Exposure Summary "Exposure Maturity Schedule" tab.

Used by the WAL (weighted average life) calculation. The real sheet lays out one
or more product blocks side by side (e.g. "NISA TIPS", "NISA SYNTHETIC US
TREASURIES", "NISA LONG TREASURIES", "NISA GOVT REPO"). Each block has maturity
("Beginning") dates down a column and value columns (Reverse Repo / Repo / Total
Return Swaps / Total), ending in a "Total" summary row.

WAL is reported on the **TIPS** block only. Per the 2026 allocation change, TIPS
was wound down and eliminated at end-June 2026 (replaced by Synthetic US
Treasuries, which are NOT included in the WAL). While TIPS exists it is the
leftmost block; once it is gone the tab has no TIPS block and WAL is reported as
zero. This parser therefore selects the block whose label contains ``TIPS`` and
returns an empty schedule (no rows) when no such block is present.

Block geometry (consistent across the NISA files): the maturity-date column sits
one column to the LEFT of the block label, and the block's "Total" column is the
one headed "Total" within the block's own column span (label_col .. label_col+5).
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from counter_risk.normalize import canonicalize_name
from counter_risk.parsers._xlsx_reader import coerce_accounting_float

_TARGET_SHEET_NAME = "Exposure Maturity Schedule"
_DEFAULT_BLOCK_KEYWORD = "TIPS"


class ExposureMaturityScheduleError(ValueError):
    """Base error for exposure maturity schedule parsing failures."""


class ExposureMaturityWorkbookLoadError(ExposureMaturityScheduleError):
    """Raised when the workbook cannot be opened/loaded."""


class ExposureMaturityWorksheetMissingError(ExposureMaturityScheduleError):
    """Raised when the required worksheet is missing."""


class ExposureMaturityColumnsMissingError(ExposureMaturityScheduleError):
    """Raised when the selected block's date/Total columns cannot be located."""


@dataclass(frozen=True)
class ExposureMaturityRow:
    """A single maturity row of the selected block."""

    maturity_date: date
    total: float


@dataclass(frozen=True)
class ExposureMaturitySchedule:
    """Parsed block schedule needed for the WAL calculation.

    ``rows`` is empty when the requested block (default: TIPS) is not present in
    the sheet -- e.g. after TIPS was eliminated -- in which case WAL is zero.
    """

    px_date: date | None
    block: str
    rows: tuple[ExposureMaturityRow, ...]


def parse_exposure_maturity_schedule(
    path: Path | str, block_keyword: str = _DEFAULT_BLOCK_KEYWORD
) -> ExposureMaturitySchedule:
    """Parse the block whose label contains ``block_keyword`` (default: TIPS).

    Returns a schedule with empty ``rows`` (not an error) when no matching block is
    present -- the expected case once TIPS has been wound down.
    """

    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Exposure maturity workbook not found: {workbook_path}")
    if workbook_path.suffix.lower() != ".xlsx":
        raise ValueError(f"Exposure maturity workbook must be an .xlsx file: {workbook_path}")

    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:  # pragma: no cover - environment dependent
        raise RuntimeError("openpyxl is required to parse exposure maturity workbooks") from exc

    try:
        workbook = load_workbook(filename=workbook_path, read_only=False, data_only=True)
    except (OSError, BadZipFile, ValueError, TypeError) as exc:
        raise ExposureMaturityWorkbookLoadError(
            f"Unable to open exposure maturity workbook: {workbook_path}"
        ) from exc

    try:
        if _TARGET_SHEET_NAME not in workbook.sheetnames:
            raise ExposureMaturityWorksheetMissingError(
                f"Missing required worksheet {_TARGET_SHEET_NAME!r} in workbook: {workbook_path}"
            )
        worksheet = workbook[_TARGET_SHEET_NAME]
        px_date = _find_px_date(worksheet)
        label_cell = _find_block_label_cell(worksheet, block_keyword)
        if label_cell is None:
            # No matching block (e.g. TIPS wound down). WAL == 0; not an error.
            return ExposureMaturitySchedule(px_date=px_date, block="", rows=())
        label_row, label_col, block_text = label_cell
        date_col, total_col = _find_block_columns(worksheet, label_row, label_col)
        rows = _parse_block_rows(
            worksheet=worksheet,
            date_col=date_col,
            total_col=total_col,
            start_row=label_row,
        )
    finally:
        workbook.close()

    return ExposureMaturitySchedule(px_date=px_date, block=block_text, rows=tuple(rows))


def _cell(worksheet: Any, row: int, col: int) -> Any:
    return worksheet.cell(row=row, column=col).value


def _as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _text(value: Any) -> str:
    return " ".join(str(value or "").split()).strip()


def _normalize_text(value: Any) -> str:
    """Preserve the parser normalization helper's established contract."""

    return canonicalize_name(str(value or ""))


def _find_px_date(worksheet: Any) -> date | None:
    max_row = int(getattr(worksheet, "max_row", 0) or 0)
    max_col = int(getattr(worksheet, "max_column", 0) or 0)
    for row in range(1, min(max_row, 30) + 1):
        for col in range(1, max_col + 1):
            if _text(_cell(worksheet, row, col)).casefold() == "px date":
                for adj in range(col + 1, min(col + 5, max_col) + 1):
                    found = _as_date(_cell(worksheet, row, adj))
                    if found is not None:
                        return found
    return None


def _find_block_label_cell(worksheet: Any, keyword: str) -> tuple[int, int, str] | None:
    """Locate the (row, col, text) of the leftmost/topmost 'NISA ...' block whose
    label contains ``keyword`` (case-insensitive). Returns None when absent."""
    kw = keyword.strip().upper()
    max_row = int(getattr(worksheet, "max_row", 0) or 0)
    max_col = int(getattr(worksheet, "max_column", 0) or 0)
    # Prefer the leftmost, then topmost, matching block.
    best_sort: tuple[int, int] | None = None
    best: tuple[int, int, str] | None = None
    for row in range(1, min(max_row, 20) + 1):
        for col in range(1, min(max_col, 40) + 1):
            text = _text(_cell(worksheet, row, col))
            upper = text.upper()
            if (
                upper.startswith("NISA ")
                and kw in upper
                and (best_sort is None or (col, row) < best_sort)
            ):
                best_sort = (col, row)
                best = (row, col, text)
    return best


def _find_block_columns(worksheet: Any, label_row: int, label_col: int) -> tuple[int, int]:
    """Return (date_col, total_col) for the block whose label is at (label_row,
    label_col).

    The maturity-date column is one column left of the label. The Total column is
    the one headed exactly "Total" within the block's own column span
    (label_col .. label_col+5); falls back to label_col+3 (the observed layout:
    Reverse Repo, Repo, Total Return Swaps, Total).
    """
    date_col = max(1, label_col - 1)

    max_row = int(getattr(worksheet, "max_row", 0) or 0)
    span_end = label_col + 5
    total_col: int | None = None
    for row in range(label_row, min(label_row + 4, max_row) + 1):
        for col in range(label_col, span_end + 1):
            if _text(_cell(worksheet, row, col)).casefold() == "total":
                total_col = col
                break
        if total_col is not None:
            break
    if total_col is None:
        total_col = label_col + 3

    # Sanity: the date column should actually carry dates below the header.
    date_hits = sum(
        1
        for row in range(label_row, min(label_row + 30, max_row) + 1)
        if _as_date(_cell(worksheet, row, date_col)) is not None
    )
    if date_hits < 1:
        raise ExposureMaturityColumnsMissingError(
            f"No maturity dates found in column {date_col} for block at row {label_row}"
        )
    return date_col, total_col


def _parse_block_rows(
    *, worksheet: Any, date_col: int, total_col: int, start_row: int
) -> list[ExposureMaturityRow]:
    """Collect (maturity_date, total) rows for the block, scanning the date column
    from ``start_row`` down.

    A "Total" summary row (or a run of empty date cells) ends the block. Blank
    Total cells are treated as zero (per the procedures' "add zeros where there
    are blanks").
    """
    max_row = int(getattr(worksheet, "max_row", 0) or 0)
    rows: list[ExposureMaturityRow] = []
    started = False
    blanks = 0
    for row in range(start_row, max_row + 1):
        cell_value = _cell(worksheet, row, date_col)
        maturity = _as_date(cell_value)
        if maturity is None:
            if _text(cell_value).casefold() == "total" and started:
                break
            if started:
                blanks += 1
                if blanks >= 3:
                    break
            continue
        blanks = 0
        started = True
        raw_total = _cell(worksheet, row, total_col)
        try:
            total = (
                coerce_accounting_float(raw_total, strip_percent=False)
                if raw_total not in (None, "")
                else 0.0
            )
        except (ValueError, TypeError):
            total = 0.0
        rows.append(ExposureMaturityRow(maturity_date=maturity, total=total))
    return rows
