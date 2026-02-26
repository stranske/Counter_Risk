"""Range-level Excel comparison helpers for regression tests."""

from __future__ import annotations

from dataclasses import dataclass
from numbers import Real
from pathlib import Path
from typing import Any


@dataclass(frozen=True)
class WorkbookRangeComparison:
    """Single range-level workbook comparison contract."""

    reference_sheet: str
    generated_sheet: str
    cell_range: str


def compare_workbook_ranges(
    reference_workbook: Path | str,
    generated_workbook: Path | str,
    *,
    comparisons: tuple[WorkbookRangeComparison, ...],
    numeric_tolerance: float = 0.0,
) -> list[str]:
    """Return range-level differences between reference and generated workbooks."""

    try:
        from openpyxl import load_workbook
        from openpyxl.utils.cell import get_column_letter, range_boundaries
    except Exception as exc:  # pragma: no cover - import guard
        raise RuntimeError("openpyxl is required for range-level workbook comparison") from exc

    reference_path = Path(reference_workbook)
    generated_path = Path(generated_workbook)
    reference = load_workbook(reference_path, data_only=True, read_only=True)
    generated = load_workbook(generated_path, data_only=True, read_only=True)
    differences: list[str] = []
    try:
        for comparison in comparisons:
            if comparison.reference_sheet not in reference.sheetnames:
                raise ValueError(
                    f"Reference workbook missing worksheet {comparison.reference_sheet!r}"
                )
            if comparison.generated_sheet not in generated.sheetnames:
                raise ValueError(
                    f"Generated workbook missing worksheet {comparison.generated_sheet!r}"
                )

            reference_sheet = reference[comparison.reference_sheet]
            generated_sheet = generated[comparison.generated_sheet]
            min_col, min_row, max_col, max_row = range_boundaries(comparison.cell_range)

            for row_number in range(min_row, max_row + 1):
                for column_number in range(min_col, max_col + 1):
                    cell_ref = f"{get_column_letter(column_number)}{row_number}"
                    reference_value = reference_sheet.cell(
                        row=row_number, column=column_number
                    ).value
                    generated_value = generated_sheet.cell(
                        row=row_number, column=column_number
                    ).value
                    if _values_equal(
                        reference_value,
                        generated_value,
                        numeric_tolerance=numeric_tolerance,
                    ):
                        continue
                    differences.append(
                        f"{comparison.reference_sheet}!{cell_ref}: "
                        f"reference={reference_value!r}, generated={generated_value!r}"
                    )
    finally:
        reference.close()
        generated.close()

    return differences


def assert_workbook_ranges_equal(
    reference_workbook: Path | str,
    generated_workbook: Path | str,
    *,
    comparisons: tuple[WorkbookRangeComparison, ...],
    numeric_tolerance: float = 0.0,
) -> None:
    """Assert that all configured cell ranges match between two workbooks."""

    differences = compare_workbook_ranges(
        reference_workbook,
        generated_workbook,
        comparisons=comparisons,
        numeric_tolerance=numeric_tolerance,
    )
    if differences:
        raise AssertionError("\n".join(differences))


def _values_equal(reference_value: Any, generated_value: Any, *, numeric_tolerance: float) -> bool:
    if isinstance(reference_value, Real) and isinstance(generated_value, Real):
        return abs(float(reference_value) - float(generated_value)) <= numeric_tolerance
    return reference_value == generated_value
