"""MOSERS futures-detail workbook write-back using openpyxl.

This module provides functions to:

1. Load an existing MOSERS workbook template while preserving all non-modified content.
2. Locate the *futures detail* section within the workbook by scanning for a
   recognised header marker cell.
3. Map matched delta-computation rows to their workbook row positions using the
   raw ``Description`` field (exact text match, no normalisation).
4. Write prior-month notional values into the designated column for each matched row.
5. Save the modified workbook back to disk, leaving all other cells and worksheets
   unchanged.

If the futures detail section cannot be located the module raises
:class:`FuturesDetailNotFoundError` immediately and no partial workbook is written.
"""

from __future__ import annotations

import logging
import os
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from counter_risk.compute.futures_delta import normalize_description
from counter_risk.io.errors import DuplicateDescriptionError

_LOG = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Cell text (case-insensitive strip match) that marks the start of the
# futures detail block.
_FUTURES_DETAIL_MARKER = "futures detail"

# Column header strings (case-insensitive strip match) that identify relevant
# columns in the table header row immediately below the marker.
_DESCRIPTION_HEADER = "description"
_PRIOR_MONTH_HEADER = "prior month notional"


# ---------------------------------------------------------------------------
# Public exceptions
# ---------------------------------------------------------------------------


class FuturesDetailNotFoundError(RuntimeError):
    """Raised when the futures detail section cannot be located in the workbook.

    This is a deterministic failure: no partial workbook is written when this
    exception propagates to the caller.
    """


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class FuturesDetailSection:
    """Describes the position of the futures detail table within a worksheet.

    Attributes
    ----------
    sheet_name:
        Name of the worksheet that contains the section.
    header_col_row:
        1-based row index of the column-header row (the row containing
        "Description", "Prior Month Notional", etc.).
    data_start_row:
        1-based row index of the first data row (row immediately below the
        column headers).
    data_end_row:
        1-based row index of the last data row (inclusive).  Rows between
        *data_start_row* and *data_end_row* (inclusive) are considered part of
        the table even if some cells are blank.
    description_col:
        1-based column index of the ``Description`` column.
    prior_month_col:
        1-based column index of the ``Prior Month Notional`` column.
    """

    sheet_name: str
    header_col_row: int
    data_start_row: int
    data_end_row: int
    description_col: int
    prior_month_col: int


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def load_mosers_workbook(path: Path | str) -> Any:
    """Load a MOSERS workbook template using openpyxl, preserving all existing content.

    Parameters
    ----------
    path:
        Path to the ``.xlsx`` workbook file.

    Returns
    -------
    openpyxl.Workbook
        The loaded workbook object.  All worksheets and cell values are
        preserved; only cells explicitly written by subsequent calls to
        :func:`write_prior_month_notional` will change.

    Raises
    ------
    FileNotFoundError
        If *path* does not exist.
    RuntimeError
        If openpyxl is not installed.
    """
    try:
        import openpyxl  # type: ignore[import-untyped]
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "openpyxl is required for MOSERS workbook write-back.  "
            "Install it with: pip install openpyxl"
        ) from exc

    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"MOSERS workbook not found: {workbook_path}")

    _LOG.debug("Loading MOSERS workbook: %s", workbook_path)
    # keep_vba=False is fine; data_only=False preserves formula strings.
    return openpyxl.load_workbook(workbook_path, data_only=False)


def locate_futures_detail_section(workbook: Any) -> FuturesDetailSection:
    """Scan *workbook* for the futures detail section and return its coordinates.

    The function iterates over every worksheet in document order, scanning each
    cell for a value that matches ``_FUTURES_DETAIL_MARKER`` (case-insensitive,
    after stripping whitespace).  When found it expects:

    * The **column-header row** to be the next non-empty row below the marker.
    * That row to contain both a ``"Description"`` column and a
      ``"Prior Month Notional"`` column (case-insensitive, after stripping).
    * One or more **data rows** immediately below the column headers.

    Data rows continue until the first row where the description cell is empty or
    the end of used rows is reached.

    Parameters
    ----------
    workbook:
        An openpyxl :class:`~openpyxl.Workbook` instance (as returned by
        :func:`load_mosers_workbook`).

    Returns
    -------
    FuturesDetailSection
        Location descriptor for the futures detail table.

    Raises
    ------
    FuturesDetailNotFoundError
        If no marker cell is found in any worksheet, if the column-header row
        cannot be identified, if required columns are missing, or if no data
        rows exist below the headers.
    """
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        section = _find_section_in_sheet(ws, sheet_name)
        if section is not None:
            return section

    raise FuturesDetailNotFoundError(
        f"Futures detail section not found in any worksheet.  "
        f"Expected a cell containing {_FUTURES_DETAIL_MARKER!r} (case-insensitive) "
        f"followed by a header row with 'Description' and 'Prior Month Notional' columns."
    )


def write_prior_month_notional(
    workbook: Any,
    section: FuturesDetailSection,
    rows: list[dict[str, Any]],
) -> int:
    """Write prior-month notional values into the workbook for matched rows.

    Each entry in *rows* is matched to a workbook data row by comparing the
    entry's ``"description"`` value to the raw text in the workbook's
    ``Description`` column (exact string match after stripping).  When a match
    is found the prior-month notional cell for that workbook row is overwritten
    with the numeric ``"prior_notional"`` value.

    Only cells within the futures detail section's prior-month notional column
    are modified.  All other cells in all worksheets remain unchanged.

    Parameters
    ----------
    workbook:
        An openpyxl :class:`~openpyxl.Workbook` instance (as returned by
        :func:`load_mosers_workbook`).
    section:
        Location descriptor returned by :func:`locate_futures_detail_section`.
    rows:
        List of result dicts from :func:`~counter_risk.compute.futures_delta.compute_futures_delta`.
        Each dict must have at least ``"description"`` and ``"prior_notional"`` keys.

    Returns
    -------
    int
        Number of workbook rows that were updated.
    """
    ws = workbook[section.sheet_name]

    # Build a lookup from raw description text → prior_notional numeric value,
    # failing fast when duplicate normalized keys are present.
    _raise_on_duplicate_normalized_descriptions(rows)
    desc_to_notional: dict[str, float] = {}
    for row in rows:
        desc = str(row.get("description", "") or "").strip()
        if desc:
            desc_to_notional[desc] = float(row.get("prior_notional", 0.0) or 0.0)

    updated = 0
    for data_row in range(section.data_start_row, section.data_end_row + 1):
        desc_cell = ws.cell(row=data_row, column=section.description_col)
        wb_desc = str(desc_cell.value or "").strip()
        if wb_desc in desc_to_notional:
            prior_cell = ws.cell(row=data_row, column=section.prior_month_col)
            prior_cell.value = desc_to_notional[wb_desc]
            updated += 1
            _LOG.debug(
                "Wrote prior_notional=%.2f for %r at row %d",
                desc_to_notional[wb_desc],
                wb_desc,
                data_row,
            )

    _LOG.info("write_prior_month_notional: updated %d/%d workbook rows", updated, len(rows))
    return updated


def save_mosers_workbook(workbook: Any, path: Path | str) -> Path:
    """Save *workbook* to *path*, creating parent directories if needed.

    All worksheets and cells that were not explicitly modified are preserved
    exactly as they were loaded.

    Parameters
    ----------
    workbook:
        An openpyxl :class:`~openpyxl.Workbook` instance.
    path:
        Destination ``.xlsx`` file path.

    Returns
    -------
    Path
        The resolved path to the saved file.
    """
    dest = Path(path)
    dest.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(dest)
    _LOG.info("Saved MOSERS workbook to: %s", dest)
    return dest


def writeback_prior_month_notionals(
    *,
    source_path: Path | str,
    output_path: Path | str,
    rows: list[dict[str, Any]],
) -> Path:
    """Write prior-month notionals to a workbook using the atomic write-back flow.

    This function is the public orchestration entrypoint for workbook write-back.
    It always routes through :func:`atomic_writeback_with_section_locate` so section
    discovery happens before any output artifact is finalized.
    """
    return atomic_writeback_with_section_locate(
        source_path=source_path,
        output_path=output_path,
        rows=rows,
    )


def atomic_writeback_with_section_locate(
    *,
    source_path: Path | str,
    output_path: Path | str,
    rows: list[dict[str, Any]],
) -> Path:
    """Atomically write prior-month notionals after locating the futures section.

    The function stages output to a temporary file in the destination directory
    and replaces *output_path* only after all steps succeed:

    1. Load source workbook.
    2. Locate futures detail section (may raise :class:`FuturesDetailNotFoundError`).
    3. Apply prior-month notional write-back.
    4. Save to temporary file.
    5. Atomically replace destination with the staged file.

    Any temporary file created for the attempt is removed in ``finally`` when
    an exception occurs, including section-location failures.
    """
    src = Path(source_path)
    dest = Path(output_path)
    dest.parent.mkdir(parents=True, exist_ok=True)

    fd, tmp_name = tempfile.mkstemp(
        prefix=f".{dest.stem}.",
        suffix=".tmp.xlsx",
        dir=str(dest.parent),
    )
    os.close(fd)
    temp_path = Path(tmp_name)

    workbook: Any | None = None
    try:
        workbook = load_mosers_workbook(src)
        section = locate_futures_detail_section(workbook)
        write_prior_month_notional(workbook, section, rows)
        save_mosers_workbook(workbook, temp_path)
        temp_path.replace(dest)
        _LOG.info("Atomic write-back complete: %s", dest)
        return dest
    finally:
        if workbook is not None and hasattr(workbook, "close"):
            workbook.close()
        if temp_path.exists():
            temp_path.unlink()


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _find_section_in_sheet(ws: Any, sheet_name: str) -> FuturesDetailSection | None:
    """Search *ws* for the futures detail marker and return the section, or None."""
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    marker_row: int | None = None
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell_val = ws.cell(row=r, column=c).value
            if cell_val is not None and str(cell_val).strip().lower() == _FUTURES_DETAIL_MARKER:
                marker_row = r
                break
        if marker_row is not None:
            break

    if marker_row is None:
        return None

    # Find the column-header row: the next row below the marker that has
    # at least one non-empty cell.
    header_row: int | None = None
    for r in range(marker_row + 1, max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        if any(v is not None and str(v).strip() != "" for v in row_vals):
            header_row = r
            break

    if header_row is None:
        raise FuturesDetailNotFoundError(
            f"Sheet {sheet_name!r}: found futures detail marker at row {marker_row} "
            f"but no column-header row follows it."
        )

    # Identify required column positions.
    description_col: int | None = None
    prior_month_col: int | None = None
    for c in range(1, max_col + 1):
        cell_val = ws.cell(row=header_row, column=c).value
        if cell_val is None:
            continue
        normalized = str(cell_val).strip().lower()
        if normalized == _DESCRIPTION_HEADER:
            description_col = c
        elif normalized == _PRIOR_MONTH_HEADER:
            prior_month_col = c

    if description_col is None:
        raise FuturesDetailNotFoundError(
            f"Sheet {sheet_name!r}: column-header row {header_row} has no 'Description' column."
        )
    if prior_month_col is None:
        raise FuturesDetailNotFoundError(
            f"Sheet {sheet_name!r}: column-header row {header_row} has no "
            f"'Prior Month Notional' column."
        )

    # Find data rows: contiguous rows starting at header_row + 1 where the
    # description cell is non-empty.
    data_start_row = header_row + 1
    if data_start_row > max_row:
        raise FuturesDetailNotFoundError(
            f"Sheet {sheet_name!r}: no data rows found below column headers at row {header_row}."
        )

    # Scan to find the last data row (stop at first empty description cell).
    data_end_row = data_start_row - 1
    for r in range(data_start_row, max_row + 1):
        desc_val = ws.cell(row=r, column=description_col).value
        if desc_val is not None and str(desc_val).strip() != "":
            data_end_row = r
        else:
            break

    if data_end_row < data_start_row:
        raise FuturesDetailNotFoundError(
            f"Sheet {sheet_name!r}: no data rows (non-empty Description cells) found "
            f"below column headers at row {header_row}."
        )

    return FuturesDetailSection(
        sheet_name=sheet_name,
        header_col_row=header_row,
        data_start_row=data_start_row,
        data_end_row=data_end_row,
        description_col=description_col,
        prior_month_col=prior_month_col,
    )


def _raise_on_duplicate_normalized_descriptions(rows: list[dict[str, Any]]) -> None:
    """Raise DuplicateDescriptionError when rows share a normalized Description."""
    normalized_to_indices: dict[str, list[int]] = {}
    for row_idx, row in enumerate(rows):
        desc = str(row.get("description", "") or "").strip()
        if not desc:
            continue
        key = normalize_description(desc)
        normalized_to_indices.setdefault(key, []).append(row_idx)

    for key, row_indices in normalized_to_indices.items():
        if len(row_indices) > 1:
            raise DuplicateDescriptionError(duplicate_key=key, row_indices=row_indices)
