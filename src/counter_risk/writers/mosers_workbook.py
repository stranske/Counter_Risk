"""Generate MOSERS-format workbooks from raw NISA monthly input."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

from counter_risk.parsers.nisa_all_programs import NisaAllProgramsData, parse_nisa_all_programs

_DISPLAY_SEGMENT = {
    "swaps": "Swaps",
    "repo": "Repo",
    "futures_cdx": "Futures / CDX",
    "futures": "Futures",
}


def generate_mosers_workbook(
    *,
    raw_nisa_path: str | Path,
    output_path: str | Path,
    as_of_date: date | datetime | None = None,
) -> Path:
    """Create a new MOSERS workbook from raw NISA All Programs input."""

    destination = Path(output_path)
    if destination.suffix.lower() != ".xlsx":
        raise ValueError(f"output_path must point to an .xlsx file: {destination}")

    parsed = parse_nisa_all_programs(raw_nisa_path)
    report_date = _to_report_date(as_of_date)

    try:
        from openpyxl import Workbook  # type: ignore[import-untyped]
    except ModuleNotFoundError as exc:  # pragma: no cover - environment dependent
        raise RuntimeError("openpyxl is required to generate MOSERS workbooks") from exc

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    ch_sheet = workbook.create_sheet(title="CPRS - CH")
    _write_ch_sheet(ch_sheet, parsed, report_date=report_date)

    fcm_sheet = workbook.create_sheet(title="CPRS - FCM")
    _write_fcm_sheet(fcm_sheet, parsed, report_date=report_date)

    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    workbook.close()
    return destination


def _to_report_date(as_of_date: date | datetime | None) -> date | datetime | str:
    if as_of_date is None:
        return "As Of Date"
    return as_of_date


def _write_ch_sheet(
    sheet: Any,
    parsed: NisaAllProgramsData,
    *,
    report_date: date | datetime | str,
) -> None:
    sheet.cell(row=1, column=1).value = "Counterparty Risk Summary"
    sheet.cell(row=2, column=1).value = f"Futures - Clearing House ({report_date})"

    sheet.cell(row=3, column=2).value = "Counterparty/ Clearing House"
    sheet.cell(row=3, column=3).value = "Cash"
    sheet.cell(row=3, column=4).value = "TIPS"
    sheet.cell(row=3, column=5).value = "Treasury"
    sheet.cell(row=3, column=6).value = "Equity"
    sheet.cell(row=3, column=7).value = "Commodity"
    sheet.cell(row=3, column=8).value = "Currency"
    sheet.cell(row=3, column=9).value = "Notional"
    sheet.cell(row=3, column=10).value = "Notional Change From Prior Month"
    sheet.cell(row=3, column=11).value = "Annualized Volatility"

    row_index = 5
    last_segment = ""
    for row in parsed.ch_rows:
        if row.segment != last_segment:
            sheet.cell(row=row_index, column=1).value = _DISPLAY_SEGMENT.get(
                row.segment, row.segment
            )
            last_segment = row.segment
            row_index += 1
        sheet.cell(row=row_index, column=2).value = row.counterparty
        sheet.cell(row=row_index, column=3).value = row.cash
        sheet.cell(row=row_index, column=4).value = row.tips
        sheet.cell(row=row_index, column=5).value = row.treasury
        sheet.cell(row=row_index, column=6).value = row.equity
        sheet.cell(row=row_index, column=7).value = row.commodity
        sheet.cell(row=row_index, column=8).value = row.currency
        sheet.cell(row=row_index, column=9).value = row.notional
        sheet.cell(row=row_index, column=10).value = row.notional_change
        sheet.cell(row=row_index, column=11).value = row.annualized_volatility
        row_index += 1


def _write_fcm_sheet(
    sheet: Any,
    parsed: NisaAllProgramsData,
    *,
    report_date: date | datetime | str,
) -> None:
    sheet.cell(row=1, column=1).value = "Counterparty Risk Summary"
    sheet.cell(row=2, column=1).value = "Futures - FCM"

    sheet.cell(row=4, column=3).value = "Counterparty/ FCM"
    sheet.cell(row=4, column=6).value = "Nominal"
    sheet.cell(row=4, column=11).value = report_date
    sheet.cell(row=4, column=12).value = "Notional change"
    sheet.cell(row=4, column=14).value = "Annualized Volatility"

    sheet.cell(row=5, column=5).value = "TIPS"
    sheet.cell(row=5, column=6).value = "Treasury"
    sheet.cell(row=5, column=7).value = "Equity"
    sheet.cell(row=5, column=8).value = "Commodity"
    sheet.cell(row=5, column=9).value = "Currency"
    sheet.cell(row=5, column=11).value = "Notional"
    sheet.cell(row=5, column=12).value = "from prior month"
    sheet.cell(row=5, column=14).value = "%"

    marker_row = 6
    sheet.cell(row=marker_row, column=3).value = "Total by Counterparty/ FCM"
    row_index = marker_row + 1
    for row in parsed.totals_rows:
        sheet.cell(row=row_index, column=3).value = row.counterparty
        sheet.cell(row=row_index, column=5).value = row.tips
        sheet.cell(row=row_index, column=6).value = row.treasury
        sheet.cell(row=row_index, column=7).value = row.equity
        sheet.cell(row=row_index, column=8).value = row.commodity
        sheet.cell(row=row_index, column=9).value = row.currency
        sheet.cell(row=row_index, column=11).value = row.notional
        sheet.cell(row=row_index, column=12).value = row.notional_change
        sheet.cell(row=row_index, column=14).value = row.annualized_volatility
        row_index += 1

    futures_marker_row = row_index + 1
    sheet.cell(row=futures_marker_row, column=3).value = "Futures Detail"
    sheet.cell(row=futures_marker_row + 1, column=3).value = "Account"
    sheet.cell(row=futures_marker_row + 1, column=5).value = "Description"
    sheet.cell(row=futures_marker_row + 1, column=7).value = "Class"
    sheet.cell(row=futures_marker_row + 1, column=8).value = "FCM"
    sheet.cell(row=futures_marker_row + 1, column=9).value = "Clearing House"
    sheet.cell(row=futures_marker_row + 1, column=12).value = "Notional"
    sheet.cell(row=futures_marker_row + 2, column=3).value = "Risk exclusive of the trend positions"
