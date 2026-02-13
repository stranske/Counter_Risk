"""Drop-in template population helpers.

This module owns the Excel template write path used to generate operator-facing
Drop-In workbook outputs.
"""

from __future__ import annotations

import re
from collections.abc import Iterable, Mapping, Sequence
from pathlib import Path
from typing import Any, cast

from counter_risk.normalize import normalize_clearing_house, normalize_counterparty

_COUNTERPARTY_COLUMNS = (
    "counterparty",
    "counterparty_name",
    "name",
    "clearing_house",
    "clearing_house_name",
)

_TEMPLATE_NUMERIC_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "cash": ("cash",),
    "tips": ("tips",),
    "treasury": ("treasury",),
    "equity": ("equity",),
    "commodity": ("commodity",),
    "currency": ("currency",),
    "notional": ("notional", "total", "total_notional"),
    "notional_change": (
        "notional_change",
        "from_prior_month",
        "prior_month_change",
        "change_from_prior_month",
    ),
}

_TEMPLATE_HEADER_LABEL_TO_METRIC: dict[str, str] = {
    "cash": "cash",
    "tips": "tips",
    "treasury": "treasury",
    "equity": "equity",
    "commodity": "commodity",
    "currency": "currency",
    "notional": "notional",
    "from prior month": "notional_change",
}


def _normalize_label(name: str) -> str:
    """Return a deterministic normalized label for row matching."""

    cleaned = " ".join(str(name).split())
    return normalize_clearing_house(normalize_counterparty(cleaned)).casefold()


def _as_path(value: str | Path, *, field_name: str) -> Path:
    if isinstance(value, Path):
        return value
    if isinstance(value, str) and value.strip():
        return Path(value)
    msg = f"{field_name} must be a non-empty path-like string or Path"
    raise ValueError(msg)


def _validate_workbook_path(path: Path, *, field_name: str, must_exist: bool) -> None:
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"{field_name} must point to an .xlsx file: {path}")

    if must_exist:
        if not path.exists():
            raise FileNotFoundError(f"Template workbook not found: {path}")
        if not path.is_file():
            raise ValueError(f"{field_name} must point to a file: {path}")
    elif path.exists() and path.is_dir():
        raise ValueError(f"{field_name} must point to a file path, not a directory: {path}")


def _coerce_breakdown(breakdown: Mapping[str, Any]) -> dict[str, float]:
    if not isinstance(breakdown, Mapping):
        raise TypeError("breakdown must be a mapping of metric names to numeric values")

    normalized: dict[str, float] = {}
    for key, raw_value in breakdown.items():
        if not isinstance(key, str) or not key.strip():
            raise ValueError("breakdown keys must be non-empty strings")

        try:
            value = float(raw_value)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"breakdown value for {key!r} must be numeric") from exc

        normalized[key] = value

    return normalized


def _is_dataframe_like(value: Any) -> bool:
    # Avoid hard dependency on pandas; this is sufficient for runtime validation.
    return hasattr(value, "columns") and hasattr(value, "to_dict")


def _iter_rows(exposures_df: Any) -> list[Mapping[str, Any]]:
    if _is_dataframe_like(exposures_df):
        rows = list(exposures_df.to_dict(orient="records"))
    elif isinstance(exposures_df, Iterable) and not isinstance(exposures_df, (str, bytes)):
        rows = list(exposures_df)
    else:
        raise TypeError(
            "exposures_df must be a pandas-like DataFrame or an iterable of row mappings"
        )

    if not rows:
        return []

    for index, row in enumerate(rows):
        if not isinstance(row, Mapping):
            raise TypeError(f"exposures_df row at index {index} must be a mapping")

    return cast(list[Mapping[str, Any]], rows)


def _build_exposure_index(rows: Sequence[Mapping[str, Any]]) -> dict[str, Mapping[str, Any]]:
    """Index rows by normalized counterparty/clearing-house label."""

    indexed: dict[str, Mapping[str, Any]] = {}
    for row in rows:
        label = None
        for key in _COUNTERPARTY_COLUMNS:
            raw_name = row.get(key)
            if isinstance(raw_name, str) and raw_name.strip():
                label = _normalize_label(raw_name)
                break

        if label is None:
            continue

        indexed[label] = row

    return indexed


def _normalize_header_label(value: str) -> str:
    collapsed = " ".join(value.split())
    return re.sub(r"[^a-z0-9]+", " ", collapsed.casefold()).strip()


def _normalize_field_name(value: str) -> str:
    collapsed = " ".join(value.split())
    return re.sub(r"[^a-z0-9]+", "_", collapsed.casefold()).strip("_")


def _find_counterparty_column(worksheet: Any, *, header_scan_rows: int = 20) -> int:
    max_row = min(getattr(worksheet, "max_row", header_scan_rows), header_scan_rows)
    max_col = min(getattr(worksheet, "max_column", 40), 40)

    for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            label = _normalize_header_label(cell.value)
            if "counterparty" in label and "clearing house" in label:
                return int(cell.column)
    return 2


def _find_numeric_template_columns(worksheet: Any, *, header_scan_rows: int = 20) -> dict[str, int]:
    max_row = min(getattr(worksheet, "max_row", header_scan_rows), header_scan_rows)
    max_col = min(getattr(worksheet, "max_column", 40), 40)

    columns: dict[str, int] = {}
    for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            label = _normalize_header_label(cell.value)
            metric = _TEMPLATE_HEADER_LABEL_TO_METRIC.get(label)
            if metric is not None:
                columns[metric] = int(cell.column)

    return columns


def _build_numeric_field_index(row: Mapping[str, Any]) -> dict[str, Any]:
    indexed: dict[str, Any] = {}
    for key, value in row.items():
        if isinstance(key, str):
            indexed[_normalize_field_name(key)] = value
    return indexed


def _coerce_numeric_cell_value(value: Any, *, field_name: str, counterparty: str) -> float:
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"Exposure value for {field_name!r} on {counterparty!r} must be numeric"
        ) from exc


def _populate_numeric_cells(
    worksheet: Any,
    exposures_by_name: Mapping[str, Mapping[str, Any]],
) -> None:
    counterparty_col = _find_counterparty_column(worksheet)
    template_numeric_columns = _find_numeric_template_columns(worksheet)
    if not template_numeric_columns:
        return

    for row_index in range(1, int(getattr(worksheet, "max_row", 0)) + 1):
        counterparty_cell = worksheet.cell(row=row_index, column=counterparty_col)
        if not isinstance(counterparty_cell.value, str) or not counterparty_cell.value.strip():
            continue

        normalized_name = _normalize_label(counterparty_cell.value)
        exposure_row = exposures_by_name.get(normalized_name)
        if exposure_row is None:
            continue

        numeric_index = _build_numeric_field_index(exposure_row)
        for metric, column in template_numeric_columns.items():
            aliases = _TEMPLATE_NUMERIC_COLUMN_ALIASES.get(metric, ())
            raw_value = None
            for alias in aliases:
                raw_value = numeric_index.get(alias)
                if raw_value is not None:
                    break

            if raw_value is None:
                continue

            worksheet.cell(row=row_index, column=column).value = _coerce_numeric_cell_value(
                raw_value,
                field_name=metric,
                counterparty=counterparty_cell.value,
            )


def fill_dropin_template(
    template_path: str | Path,
    exposures_df: Any,
    breakdown: Mapping[str, Any],
    *,
    output_path: str | Path,
) -> Path:
    """Load a drop-in template and write a populated output workbook.

    The population logic is intentionally conservative in this first slice: inputs
    are validated, row labels are indexed for deterministic matching, and the
    template is loaded/saved to guarantee a valid output workbook for follow-on
    cell population tasks.
    """

    template_file = _as_path(template_path, field_name="template_path")
    output_file = _as_path(output_path, field_name="output_path")
    _validate_workbook_path(template_file, field_name="template_path", must_exist=True)
    _validate_workbook_path(output_file, field_name="output_path", must_exist=False)

    rows = _iter_rows(exposures_df)
    exposures_by_name = _build_exposure_index(rows)
    _ = _coerce_breakdown(breakdown)

    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ModuleNotFoundError as exc:  # pragma: no cover - environment dependent
        raise RuntimeError(
            "openpyxl is required to fill drop-in templates. "
            "Install project dev dependencies to enable this feature."
        ) from exc

    try:
        workbook = load_workbook(filename=template_file)
    except OSError as exc:
        raise ValueError(f"Unable to load template workbook: {template_file}") from exc

    worksheet = workbook.active
    _populate_numeric_cells(worksheet, exposures_by_name)

    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_file)
    workbook.close()
    return output_file
