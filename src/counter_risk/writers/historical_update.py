"""Historical workbook append helpers for Counter Risk chart-linked files.

This module centralizes shared workbook/date/header validation used by the
historical graph workbook update path. Variant-specific row append functions are
implemented in follow-on slices.
"""

from __future__ import annotations

import logging
from collections.abc import Mapping
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

LOGGER = logging.getLogger(__name__)

SHEET_ALL_PROGRAMS_3_YEAR = "All Programs 3 Year"
SHEET_EX_LLC_3_YEAR = "ex LLC 3 Year"
SHEET_LLC_3_YEAR = "LLC 3 Year"
SHEET_WAL = "WAL"

SERIES_BY_SHEET: dict[str, tuple[str, ...]] = {
    SHEET_ALL_PROGRAMS_3_YEAR: (
        "Total",
        "Total x-clearing",
        "Cash",
        "Class",
        "TIPS",
        "Nominal",
        "Equity",
        "Currency",
        "Commodity",
    ),
    SHEET_EX_LLC_3_YEAR: (
        "Total",
        "Class",
        "TIPS",
        "Nominal",
        "Equity",
        "Commodity",
        "Currency",
    ),
    SHEET_LLC_3_YEAR: (
        "Total",
        "Class",
        "Nominal",
        "Equity",
        "Commodity",
        "Currency",
    ),
}

DATE_HEADER_CANDIDATES: tuple[str, ...] = ("date", "as of date", "as-of date")
WAL_HEADER_CANDIDATES: tuple[str, ...] = ("wal", "wal tips repo", "weighted average life")
HEADER_SCAN_ROWS = 12
_DEFAULT_EX_LLC_3_YEAR_RELATIVE_PATH = Path(
    "docs/Ratings Instructiosns/Historical Counterparty Risk Graphs - ex LLC 3 Year.xlsx"
)


class HistoricalUpdateError(ValueError):
    """Base error for historical workbook update failures."""


class WorkbookValidationError(HistoricalUpdateError):
    """Raised when a workbook path or structure is invalid."""


class WorksheetNotFoundError(HistoricalUpdateError):
    """Raised when a required worksheet does not exist."""


class AppendDateError(HistoricalUpdateError):
    """Raised when append date resolution or ordering checks fail."""


class AppendDateResolutionError(AppendDateError):
    """Raised when append date cannot be resolved from any supported source."""


class DateMonotonicityError(AppendDateError):
    """Raised when append date is not newer than the latest existing row date."""


@dataclass(frozen=True)
class WalSheetAppendLocation:
    """Resolved WAL sheet coordinates for a single-row append."""

    sheet_name: str
    header_row: int
    date_column: int
    wal_column: int
    append_row: int
    last_dated_row: int | None


@dataclass(frozen=True)
class _CellSnapshot:
    value: Any
    is_formula: bool
    style_id: int | None
    number_format: str | None


def _resolve_repo_root() -> Path:
    return Path(__file__).resolve().parents[3]


def _as_path(value: str | Path, *, field_name: str) -> Path:
    if isinstance(value, Path):
        return value
    if isinstance(value, str) and value.strip():
        return Path(value)
    raise WorkbookValidationError(f"{field_name} must be a non-empty path-like value")


def _validate_workbook_path(path: Path, *, field_name: str = "workbook_path") -> None:
    if path.suffix.lower() != ".xlsx":
        raise WorkbookValidationError(f"{field_name} must point to an .xlsx file: {path}")
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")
    if not path.is_file():
        raise WorkbookValidationError(f"{field_name} must point to a file: {path}")


def locate_ex_llc_3_year_workbook(
    *, search_root: str | Path | None = None, expected_relative_path: Path | None = None
) -> Path:
    """Resolve the ex LLC 3 Year workbook path under the expected repository location."""

    root = (
        _resolve_repo_root()
        if search_root is None
        else _as_path(search_root, field_name="search_root")
    )
    relative_path = expected_relative_path or _DEFAULT_EX_LLC_3_YEAR_RELATIVE_PATH
    workbook_path = root / relative_path
    _validate_workbook_path(workbook_path, field_name="hist_ex_llc_3yr_xlsx")
    return workbook_path


def open_ex_llc_3_year_workbook(
    *, search_root: str | Path | None = None, expected_relative_path: Path | None = None
) -> tuple[Path, Any]:
    """Locate and open the ex LLC 3 Year workbook at its expected path."""

    workbook_path = locate_ex_llc_3_year_workbook(
        search_root=search_root,
        expected_relative_path=expected_relative_path,
    )
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ModuleNotFoundError as exc:  # pragma: no cover - environment dependent
        raise RuntimeError(
            "openpyxl is required to open historical workbooks. "
            "Install project dev dependencies to enable this feature."
        ) from exc

    try:
        workbook = load_workbook(filename=workbook_path)
    except Exception as exc:
        raise WorkbookValidationError(f"Unable to load workbook: {workbook_path}") from exc
    return workbook_path, workbook


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split()).casefold()


def _find_header_row(
    worksheet: Any,
    *,
    max_scan_rows: int = HEADER_SCAN_ROWS,
    max_scan_cols: int = 40,
) -> int:
    upper_row = max_scan_rows
    upper_col = min(int(getattr(worksheet, "max_column", max_scan_cols)), max_scan_cols)

    for row_index in range(1, upper_row + 1):
        for col_index in range(1, upper_col + 1):
            header_value = _normalize_header(worksheet.cell(row=row_index, column=col_index).value)
            if header_value in DATE_HEADER_CANDIDATES:
                return row_index

    raise WorkbookValidationError(
        f"Unable to locate a header row containing a date label in worksheet {worksheet.title!r}"
    )


def _extract_column_header_values(
    worksheet: Any,
    *,
    column_index: int,
    max_scan_rows: int = HEADER_SCAN_ROWS,
) -> tuple[str, ...]:
    values: list[str] = []
    for row_index in range(1, max_scan_rows + 1):
        raw_value = _get_cell_value_no_create(
            worksheet,
            row=row_index,
            column=column_index,
        )
        if not isinstance(raw_value, str):
            continue
        normalized = _normalize_header(raw_value)
        if normalized:
            values.append(normalized)
    return tuple(values)


def _get_cell_value_no_create(worksheet: Any, *, row: int, column: int) -> Any:
    existing_cells = getattr(worksheet, "_cells", None)
    if isinstance(existing_cells, dict):
        existing = existing_cells.get((row, column))
        return None if existing is None else existing.value
    if hasattr(worksheet, "iter_rows"):
        for row_values in worksheet.iter_rows(
            min_row=row,
            max_row=row,
            min_col=column,
            max_col=column,
            values_only=True,
        ):
            return row_values[0]
        return None
    return worksheet.cell(row=row, column=column).value


def _build_consolidated_header_map(
    worksheet: Any,
    *,
    max_scan_rows: int = HEADER_SCAN_ROWS,
    max_scan_cols: int = 256,
) -> dict[int, str]:
    upper_col = min(int(getattr(worksheet, "max_column", max_scan_cols)), max_scan_cols)
    header_map: dict[int, str] = {}
    for col_index in range(1, upper_col + 1):
        stacked_values = _extract_column_header_values(
            worksheet,
            column_index=col_index,
            max_scan_rows=max_scan_rows,
        )
        if not stacked_values:
            continue
        header_map[col_index] = " ".join(stacked_values)
    return header_map


def _build_header_map(
    worksheet: Any,
    *,
    header_row: int,
    max_scan_cols: int = 256,
) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for col_index, normalized in _build_consolidated_header_map(
        worksheet,
        max_scan_rows=max(header_row, HEADER_SCAN_ROWS),
        max_scan_cols=max_scan_cols,
    ).items():
        header_map[normalized] = col_index
    return header_map


def _resolve_append_date(
    *,
    append_date: date | None,
    config_as_of_date: date | None,
    rollup_data: Mapping[str, Any] | None = None,
) -> date:
    if append_date is not None:
        return append_date
    if config_as_of_date is not None:
        return config_as_of_date

    inferred = _infer_date_from_cprs_ch_header(rollup_data)
    if inferred is not None:
        return inferred

    raise AppendDateResolutionError(
        "Unable to resolve append date from append_date, config_as_of_date, or CPRS-CH header date."
    )


def _infer_date_from_cprs_ch_header(rollup_data: Mapping[str, Any] | None) -> date | None:
    if not rollup_data:
        return None

    candidate_labels = {
        "cprs ch header date",
        "cprs-ch header date",
        "cprs_ch_header_date",
        "cprs ch as of date",
        "cprs_ch_as_of_date",
        "as of date",
        "as_of_date",
        "report date",
        "report_date",
    }
    for raw_key, raw_value in rollup_data.items():
        normalized_key = _normalize_header(raw_key)
        if normalized_key not in candidate_labels:
            continue

        parsed = _coerce_cell_date(raw_value)
        if parsed is not None:
            return parsed

    return None


def _coerce_rollup_data(rollup_data: Mapping[str, Any]) -> dict[str, float]:
    normalized: dict[str, float] = {}
    for raw_key, raw_value in rollup_data.items():
        key = _normalize_header(raw_key)
        if not key:
            continue
        try:
            numeric_value = float(raw_value)
        except (TypeError, ValueError) as exc:
            raise HistoricalUpdateError(f"Rollup value for {raw_key!r} must be numeric") from exc
        normalized[key] = numeric_value
    return normalized


def _get_date_column(header_map: Mapping[str, int]) -> int:
    for candidate in DATE_HEADER_CANDIDATES:
        column = header_map.get(candidate)
        if column is not None:
            return column
    raise WorkbookValidationError("Worksheet header row is missing a date column")


def _get_date_column_from_consolidated(header_map: Mapping[int, str]) -> int:
    for column_index, normalized in header_map.items():
        if normalized in DATE_HEADER_CANDIDATES:
            return column_index
    raise WorkbookValidationError("Worksheet header rows are missing a date column")


def _get_numeric_series_columns(
    header_map: Mapping[int, str], *, date_column: int
) -> dict[int, str]:
    return {
        column_index: normalized
        for column_index, normalized in header_map.items()
        if column_index != date_column
    }


def _coerce_cell_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(stripped, fmt).date()
            except ValueError:
                continue
    raise AppendDateError(f"Unable to parse existing worksheet date value: {value!r}")


def _find_last_dated_row(worksheet: Any, *, header_row: int, date_column: int) -> int | None:
    max_row = int(getattr(worksheet, "max_row", header_row))
    for row_index in range(max_row, header_row, -1):
        raw_date = worksheet.cell(row=row_index, column=date_column).value
        parsed = _coerce_cell_date(raw_date)
        if parsed is not None:
            return row_index
    return None


def _get_wal_column_from_consolidated(header_map: Mapping[int, str], *, date_column: int) -> int:
    for column_index, normalized in header_map.items():
        if column_index == date_column:
            continue
        if normalized in WAL_HEADER_CANDIDATES:
            return column_index
        if normalized.startswith("wal "):
            return column_index
    raise WorkbookValidationError("Worksheet header rows are missing a WAL value column")


def read_wal_sheet_append_location(
    workbook: Any,
    *,
    wal_sheet_name: str = SHEET_WAL,
) -> WalSheetAppendLocation:
    """Resolve WAL sheet structure and row location for appending one WAL observation."""

    if wal_sheet_name not in getattr(workbook, "sheetnames", []):
        raise WorksheetNotFoundError(f"Required worksheet not found: {wal_sheet_name}")

    worksheet = workbook[wal_sheet_name]
    header_row = _find_header_row(worksheet)
    consolidated_headers = _build_consolidated_header_map(
        worksheet,
        max_scan_rows=max(header_row, HEADER_SCAN_ROWS),
    )
    date_column = _get_date_column_from_consolidated(consolidated_headers)
    wal_column = _get_wal_column_from_consolidated(
        consolidated_headers,
        date_column=date_column,
    )
    last_dated_row = _find_last_dated_row(worksheet, header_row=header_row, date_column=date_column)
    append_row = header_row + 1 if last_dated_row is None else last_dated_row + 1
    return WalSheetAppendLocation(
        sheet_name=wal_sheet_name,
        header_row=header_row,
        date_column=date_column,
        wal_column=wal_column,
        append_row=append_row,
        last_dated_row=last_dated_row,
    )


def _snapshot_preserved_wal_cells(
    worksheet: Any,
    *,
    preserve_through_row: int,
    preserve_through_column: int,
) -> dict[tuple[int, int], _CellSnapshot]:
    snapshots: dict[tuple[int, int], _CellSnapshot] = {}
    existing_cells = getattr(worksheet, "_cells", None)
    if isinstance(existing_cells, dict):
        for (row_index, column_index), cell in existing_cells.items():
            if row_index > preserve_through_row or column_index > preserve_through_column:
                continue
            value = getattr(cell, "value", None)
            snapshots[(row_index, column_index)] = _CellSnapshot(
                value=value,
                is_formula=isinstance(value, str) and value.startswith("="),
                style_id=getattr(cell, "style_id", None),
                number_format=getattr(cell, "number_format", None),
            )
        return snapshots

    for row_index in range(1, preserve_through_row + 1):
        for column_index in range(1, preserve_through_column + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            value = cell.value
            snapshots[(row_index, column_index)] = _CellSnapshot(
                value=value,
                is_formula=isinstance(value, str) and value.startswith("="),
                style_id=getattr(cell, "style_id", None),
                number_format=getattr(cell, "number_format", None),
            )
    return snapshots


def _validate_preserved_wal_cells(
    worksheet: Any, expected_snapshots: Mapping[tuple[int, int], _CellSnapshot]
) -> None:
    for (row_index, column_index), expected in expected_snapshots.items():
        cell = worksheet.cell(row=row_index, column=column_index)
        value = cell.value
        is_formula = isinstance(value, str) and value.startswith("=")
        if expected.is_formula != is_formula:
            raise WorkbookValidationError(
                "WAL append changed existing formula usage at "
                f"row={row_index} column={column_index}"
            )
        if expected.value != value:
            raise WorkbookValidationError(
                "WAL append changed existing cell value at "
                f"row={row_index} column={column_index}"
            )
        if expected.style_id != getattr(cell, "style_id", None):
            raise WorkbookValidationError(
                "WAL append changed existing formatting at "
                f"row={row_index} column={column_index}"
            )
        if expected.number_format != getattr(cell, "number_format", None):
            raise WorkbookValidationError(
                "WAL append changed existing number format at "
                f"row={row_index} column={column_index}"
            )


def _copy_row_cell_presentation(
    worksheet: Any, *, source_row: int, target_row: int, columns: tuple[int, ...]
) -> None:
    for column_index in columns:
        source_cell = worksheet.cell(row=source_row, column=column_index)
        target_cell = worksheet.cell(row=target_row, column=column_index)
        target_cell._style = copy(source_cell._style)
        target_cell.number_format = source_cell.number_format


def _append_to_sheet(
    *,
    workbook: Any,
    sheet_name: str,
    rollup_data: Mapping[str, Any],
    resolved_date: date,
) -> None:
    if sheet_name not in getattr(workbook, "sheetnames", []):
        raise WorksheetNotFoundError(f"Required worksheet not found: {sheet_name}")

    worksheet = workbook[sheet_name]
    header_row = _find_header_row(worksheet)
    consolidated_headers = _build_consolidated_header_map(worksheet, max_scan_rows=HEADER_SCAN_ROWS)
    date_column = _get_date_column_from_consolidated(consolidated_headers)
    numeric_series_columns = _get_numeric_series_columns(
        consolidated_headers,
        date_column=date_column,
    )
    if not numeric_series_columns:
        LOGGER.warning("No numeric series columns detected in %s", sheet_name)

    last_row = _find_last_dated_row(worksheet, header_row=header_row, date_column=date_column)
    if last_row is None:
        target_row = HEADER_SCAN_ROWS + 1
    else:
        last_date = _coerce_cell_date(worksheet.cell(row=last_row, column=date_column).value)
        if last_date is None:
            raise AppendDateError(
                f"Last worksheet date is blank for sheet {sheet_name!r} at row {last_row}"
            )
        if resolved_date <= last_date:
            raise DateMonotonicityError(
                "Append date must be newer than the last row date: "
                f"append_date={resolved_date.isoformat()} last_row_date={last_date.isoformat()}"
            )
        target_row = last_row + 1

    worksheet.cell(row=target_row, column=date_column).value = resolved_date

    normalized_rollups = _coerce_rollup_data(rollup_data)
    for series_column, normalized_series in numeric_series_columns.items():
        worksheet.cell(row=target_row, column=series_column).value = normalized_rollups.get(
            normalized_series, 0.0
        )


def _append_row(
    *,
    workbook_path: str | Path,
    sheet_name: str,
    rollup_data: Mapping[str, Any],
    append_date: date | None,
    config_as_of_date: date | None,
) -> Path:
    path = _as_path(workbook_path, field_name="workbook_path")
    _validate_workbook_path(path)
    resolved_date = _resolve_append_date(
        append_date=append_date,
        config_as_of_date=config_as_of_date,
        rollup_data=rollup_data,
    )

    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:  # pragma: no cover - environment dependent
        raise RuntimeError(
            "openpyxl is required to update historical workbooks. "
            "Install project dev dependencies to enable this feature."
        ) from exc

    try:
        workbook = load_workbook(filename=path)
    except Exception as exc:
        raise WorkbookValidationError(f"Unable to load workbook: {path}") from exc

    try:
        _append_to_sheet(
            workbook=workbook,
            sheet_name=sheet_name,
            rollup_data=rollup_data,
            resolved_date=resolved_date,
        )
        workbook.save(path)
    finally:
        workbook.close()
    return path


def append_row_all_programs(
    workbook_path: str | Path,
    rollup_data: Mapping[str, Any],
    *,
    append_date: date | None = None,
    config_as_of_date: date | None = None,
) -> Path:
    """Append one row to the `All Programs 3 Year` worksheet."""
    return _append_row(
        workbook_path=workbook_path,
        sheet_name=SHEET_ALL_PROGRAMS_3_YEAR,
        rollup_data=rollup_data,
        append_date=append_date,
        config_as_of_date=config_as_of_date,
    )


def append_row_ex_trend(
    workbook_path: str | Path,
    rollup_data: Mapping[str, Any],
    *,
    append_date: date | None = None,
    config_as_of_date: date | None = None,
) -> Path:
    """Append one row to the `ex LLC 3 Year` worksheet."""
    return _append_row(
        workbook_path=workbook_path,
        sheet_name=SHEET_EX_LLC_3_YEAR,
        rollup_data=rollup_data,
        append_date=append_date,
        config_as_of_date=config_as_of_date,
    )


def append_row_trend(
    workbook_path: str | Path,
    rollup_data: Mapping[str, Any],
    *,
    append_date: date | None = None,
    config_as_of_date: date | None = None,
) -> Path:
    """Append one row to the `LLC 3 Year` worksheet."""
    return _append_row(
        workbook_path=workbook_path,
        sheet_name=SHEET_LLC_3_YEAR,
        rollup_data=rollup_data,
        append_date=append_date,
        config_as_of_date=config_as_of_date,
    )


def append_wal_row(
    workbook_path: str | Path,
    *,
    px_date: date,
    wal_value: float,
    wal_sheet_name: str = SHEET_WAL,
) -> Path:
    """Append one WAL row to the historical WAL sheet."""

    path = _as_path(workbook_path, field_name="workbook_path")
    _validate_workbook_path(path)

    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:  # pragma: no cover - environment dependent
        raise RuntimeError(
            "openpyxl is required to update historical workbooks. "
            "Install project dev dependencies to enable this feature."
        ) from exc

    workbook = None
    try:
        workbook = load_workbook(filename=path)
        append_target = read_wal_sheet_append_location(workbook, wal_sheet_name=wal_sheet_name)
        worksheet = workbook[append_target.sheet_name]
        preserve_up_to_row = append_target.last_dated_row or append_target.header_row
        preserve_snapshots = _snapshot_preserved_wal_cells(
            worksheet,
            preserve_through_row=preserve_up_to_row,
            preserve_through_column=max(
                append_target.date_column,
                append_target.wal_column,
                int(getattr(worksheet, "max_column", append_target.wal_column)),
            ),
        )

        if append_target.last_dated_row is not None:
            last_date = _coerce_cell_date(
                worksheet.cell(
                    row=append_target.last_dated_row, column=append_target.date_column
                ).value
            )
            if last_date is None:
                raise AppendDateError(
                    "Last worksheet date is blank for WAL sheet "
                    f"{append_target.sheet_name!r} at row {append_target.last_dated_row}"
                )
            if px_date <= last_date:
                raise DateMonotonicityError(
                    "Append date must be newer than the last row date: "
                    f"append_date={px_date.isoformat()} last_row_date={last_date.isoformat()}"
                )
            _copy_row_cell_presentation(
                worksheet,
                source_row=append_target.last_dated_row,
                target_row=append_target.append_row,
                columns=(append_target.date_column, append_target.wal_column),
            )

        worksheet.cell(row=append_target.append_row, column=append_target.date_column).value = (
            px_date
        )
        worksheet.cell(row=append_target.append_row, column=append_target.wal_column).value = float(
            wal_value
        )
        _validate_preserved_wal_cells(worksheet, preserve_snapshots)
        workbook.save(path)
    finally:
        if workbook is not None:
            workbook.close()
    return path


__all__ = [
    "AppendDateError",
    "AppendDateResolutionError",
    "DateMonotonicityError",
    "HistoricalUpdateError",
    "SHEET_ALL_PROGRAMS_3_YEAR",
    "SHEET_EX_LLC_3_YEAR",
    "SHEET_LLC_3_YEAR",
    "SHEET_WAL",
    "SERIES_BY_SHEET",
    "WalSheetAppendLocation",
    "WorkbookValidationError",
    "WorksheetNotFoundError",
    "append_row_all_programs",
    "append_row_ex_trend",
    "append_row_trend",
    "append_wal_row",
    "locate_ex_llc_3_year_workbook",
    "open_ex_llc_3_year_workbook",
    "read_wal_sheet_append_location",
]
